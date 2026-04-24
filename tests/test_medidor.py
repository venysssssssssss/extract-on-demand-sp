from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl

import sap_automation.medidor as medidor_module
from sap_automation.medidor import (
    MedidorExtractor,
    compact_medidor_raw_exports,
    collect_equipments_from_el31_export,
    collect_iq09_grpreg_by_equipment,
    compute_medidor_el31_period,
    load_grpreg_type_map,
    load_workbook_column,
    run_medidor_demandante,
    write_medidor_final_csv,
)
from sap_automation.medidor_repository import (
    MEDIDOR_INGEST_CHUNK_SIZE,
    MedidorRepository,
    _chunked,
    read_medidor_final_csv,
    sm_dados_medidor_sp,
)


def _write_xlsx(path: Path, headers: list[str], rows: list[list[str]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_compute_medidor_el31_period_uses_previous_year_start_to_today() -> None:
    period_from, period_to, reference = compute_medidor_el31_period(date(2026, 4, 16))

    assert period_from == "01.01.2025"
    assert period_to == "16.04.2026"
    assert reference == "20260416"


def test_load_workbook_column_deduplicates_installations(tmp_path: Path) -> None:
    workbook_path = tmp_path / "instalacaosp.xlsx"
    _write_xlsx(
        workbook_path,
        ["INSTALACAO"],
        [["ATE0000002"], ["MTE0012436"], ["ATE0000002"], [""]],
    )

    values = load_workbook_column(workbook_path, column_name="INSTALACAO")

    assert values == ["ATE0000002", "MTE0012436"]


def test_medidor_repository_fetches_alimentador_and_saves_meter_types(tmp_path: Path) -> None:
    db_url = f"sqlite+pysqlite:///{tmp_path / 'medidor.db'}"
    repository = MedidorRepository(db_url)
    from sqlalchemy import Column, MetaData, String, Table, create_engine, insert, select

    engine = create_engine(db_url)
    source = Table(
        "TBL_REINCIDENCIA_SM",
        MetaData(),
        Column("ALIMENTADOR", String(128)),
        Column("DISTRIBUIDORA", String(128)),
    )
    source.create(engine)
    with engine.begin() as conn:
        conn.execute(
            insert(source),
            [
                {"ALIMENTADOR": "ATE0000002", "DISTRIBUIDORA": "São Paulo"},
                {"ALIMENTADOR": "ATE0000002", "DISTRIBUIDORA": "São Paulo"},
                {"ALIMENTADOR": "MTE0012436", "DISTRIBUIDORA": "São Paulo"},
                {"ALIMENTADOR": "RJ0000001", "DISTRIBUIDORA": "Rio"},
            ],
        )

    installations = repository.get_installations_by_alimentador(distribuidora="São Paulo")
    rows_ingested = repository.save_meter_types(
        [
            {"instalacao": "ATE0000002", "tipo": "Digital"},
            {"instalacao": "MTE0012436", "tipo": "Analógico"},
        ]
    )

    with engine.connect() as conn:
        saved_rows = conn.execute(select(sm_dados_medidor_sp).order_by(sm_dados_medidor_sp.c.num_instalacao)).fetchall()

    assert installations == ["ATE0000002", "MTE0012436"]
    assert rows_ingested == 2
    assert [(row.num_instalacao, row.tp_medidor) for row in saved_rows] == [
        ("ATE0000002", "Digital"),
        ("MTE0012436", "Analógico"),
    ]


def test_medidor_repository_uses_2000_row_ingest_chunks() -> None:
    values = list(range(4501))

    chunks = _chunked(values, size=MEDIDOR_INGEST_CHUNK_SIZE)

    assert [len(chunk) for chunk in chunks] == [2000, 2000, 501]

def test_load_grpreg_type_map_reads_expected_columns(tmp_path: Path) -> None:
    workbook_path = tmp_path / "gruporegsap.xlsx"
    _write_xlsx(
        workbook_path,
        ["Grp.registrad.", "Tipo"],
        [["AD30002N", "Digital"], ["DA01010", "Analógico"]],
    )

    mapping = load_grpreg_type_map(workbook_path)

    assert mapping == {"AD30002N": "Digital", "DA01010": "Analógico"}


def test_collect_exports_and_write_final_classified_csv(tmp_path: Path) -> None:
    el31_path = tmp_path / "el31.txt"
    iq09_path = tmp_path / "iq09.txt"
    final_path = tmp_path / "final.csv"
    el31_path.write_text(
        "Instalação\tEquipamento\nATE0000002\tEQ001\nMTE0012436\tEQ002\n",
        encoding="cp1252",
    )
    iq09_path.write_text("Equipamento\tGrpReg.\nEQ001\tAD30002N\nEQ002\tDA01010\n", encoding="utf-8")

    el31_rows, equipments = collect_equipments_from_el31_export(el31_path)
    iq09_mapping = collect_iq09_grpreg_by_equipment([iq09_path])
    rows_written = write_medidor_final_csv(
        el31_rows=el31_rows,
        iq09_grpreg_by_equipment=iq09_mapping,
        grpreg_type_map={"AD30002N": "Digital", "DA01010": "Analógico"},
        output_path=final_path,
    )

    assert equipments == ["EQ001", "EQ002"]
    assert iq09_mapping == {"EQ001": "AD30002N", "EQ002": "DA01010"}
    assert rows_written == 2
    assert final_path.read_text(encoding="utf-8").splitlines() == [
        "instalacao,equipamento,grp_reg,tipo",
        "ATE0000002,EQ001,AD30002N,Digital",
        "MTE0012436,EQ002,DA01010,Analógico",
    ]
    assert read_medidor_final_csv(final_path) == [
        {"instalacao": "ATE0000002", "equipamento": "EQ001", "grp_reg": "AD30002N", "tipo": "Digital"},
        {"instalacao": "MTE0012436", "equipamento": "EQ002", "grp_reg": "DA01010", "tipo": "Analógico"},
    ]


def test_collect_iq09_accepts_corrupted_serial_number_header(tmp_path: Path) -> None:
    iq09_path = tmp_path / "iq09.txt"
    iq09_path.write_text(
        "N� s�rie\tStatUsu�r.\tGrpReg.\nEQ001\tINST\tAD30002N\nEQ002\tINST\tDA01010\n",
        encoding="utf-8",
    )

    mapping = collect_iq09_grpreg_by_equipment([iq09_path])

    assert mapping == {"EQ001": "AD30002N", "EQ002": "DA01010"}


def test_compact_medidor_raw_exports_deduplicates_all_raw_txt(tmp_path: Path) -> None:
    raw_dir = tmp_path / "runs" / "run-medidor" / "medidor" / "raw"
    raw_dir.mkdir(parents=True)
    output_path = tmp_path / "compactado.csv"
    (raw_dir / "el31_medidor_20260416_run-medidor_001.txt").write_text(
        "Instalação\tEquipamento\nATE0000002\tEQ001\nMTE0012436\tEQ002\n",
        encoding="utf-8",
    )
    (raw_dir / "el31_medidor_20260416_run-medidor_002.txt").write_text(
        "Instalação\tEquipamento\nATE0000002\tEQ001\nXTE9999999\tEQ003\n",
        encoding="utf-8",
    )

    result = compact_medidor_raw_exports(
        raw_dir=raw_dir,
        output_csv_path=output_path,
    )

    assert result.el31_rows_read == 4
    assert result.deduped_rows_written == 3
    assert result.duplicate_equipments_removed == 1
    assert result.iq09_raw_paths == []
    assert result.equipments_without_iq09_group == 0
    assert result.equipments_without_type == 0
    assert output_path.read_text(encoding="utf-8").splitlines() == [
        "instalacao,unid_leit,equipamento,dta_leit_pr",
        "ATE0000002,,EQ001,",
        "MTE0012436,,EQ002,",
        "XTE9999999,,EQ003,",
    ]
    assert Path(result.manifest_path).exists()


def test_compact_medidor_raw_exports_reads_sap_alv_txt_with_preamble(tmp_path: Path) -> None:
    raw_dir = tmp_path / "runs" / "run-medidor" / "medidor" / "raw"
    raw_dir.mkdir(parents=True)
    output_path = tmp_path / "compactado.csv"
    (raw_dir / "el31_medidor_20260416_run-medidor_001.txt").write_text(
        "\n".join(
            [
                "Relatorio EL31",
                "-----------------------------------------------",
                "|Instalação |Unid.leit .|Equipamento|TL |DtaLeitPr.|",
                "|ATE0000002 |001        |EQ001      |A  |16.04.2026|",
                "|MTE0012436 |002        |EQ002      |B  |16.04.2026|",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    result = compact_medidor_raw_exports(raw_dir=raw_dir, output_csv_path=output_path)

    assert result.el31_rows_read == 2
    assert result.deduped_rows_written == 2
    assert output_path.read_text(encoding="utf-8").splitlines() == [
        "instalacao,unid_leit,equipamento,dta_leit_pr",
        "ATE0000002,001,EQ001,16.04.2026",
        "MTE0012436,002,EQ002,16.04.2026",
    ]


def test_compact_medidor_raw_exports_normalizes_replacement_char_installation_header(tmp_path: Path) -> None:
    raw_dir = tmp_path / "runs" / "run-medidor" / "medidor" / "raw"
    raw_dir.mkdir(parents=True)
    output_path = tmp_path / "compactado.csv"
    (raw_dir / "el31_medidor_20260416_run-medidor_001.txt").write_text(
        "Instala��o\tUnid.leit .\tEquipamento\tTL \tDtaLeitPr.\nATE0000002\t001\tEQ001\tA\t16.04.2026\n",
        encoding="utf-8",
    )

    result = compact_medidor_raw_exports(raw_dir=raw_dir, output_csv_path=output_path)

    assert result.deduped_rows_written == 1
    assert output_path.read_text(encoding="utf-8").splitlines() == [
        "instalacao,unid_leit,equipamento,dta_leit_pr",
        "ATE0000002,001,EQ001,16.04.2026",
    ]


def test_compact_medidor_raw_exports_skips_el31_file_without_equipment_rows(tmp_path: Path) -> None:
    raw_dir = tmp_path / "runs" / "run-medidor" / "medidor" / "raw"
    raw_dir.mkdir(parents=True)
    output_path = tmp_path / "compactado.csv"
    (raw_dir / "el31_medidor_20260416_run-medidor_001.txt").write_text(
        "Instalação\tEquipamento\nATE0000002\tEQ001\n",
        encoding="utf-8",
    )
    (raw_dir / "el31_medidor_20260416_run-medidor_002.txt").write_bytes(
        b"Relat\xedrio EL31\r\nMensagem SAP sem tabela de equipamentos\r\n",
    )

    result = compact_medidor_raw_exports(raw_dir=raw_dir, output_csv_path=output_path)

    assert result.el31_rows_read == 1
    assert result.deduped_rows_written == 1
    assert len(result.el31_raw_paths_skipped) == 1
    assert "no rows with equipment found" in result.el31_raw_paths_skipped[0]
    assert output_path.read_text(encoding="utf-8").splitlines() == [
        "instalacao,unid_leit,equipamento,dta_leit_pr",
        "ATE0000002,,EQ001,",
    ]


def test_compact_medidor_raw_exports_accepts_eq_tl_as_equipment_header(tmp_path: Path) -> None:
    raw_dir = tmp_path / "runs" / "run-medidor" / "medidor" / "raw"
    raw_dir.mkdir(parents=True)
    output_path = tmp_path / "compactado.csv"
    (raw_dir / "el31_medidor_20260416_run-medidor_003.txt").write_text(
        "Instala��o\tUnid.leit .\tEq. TL\tDtaLeitPr.\nATE0000002\t001\tEQ001\t16.04.2026\n",
        encoding="utf-8",
    )

    result = compact_medidor_raw_exports(raw_dir=raw_dir, output_csv_path=output_path)

    assert result.el31_rows_read == 1
    assert result.deduped_rows_written == 1
    assert result.el31_raw_paths_skipped == []
    assert output_path.read_text(encoding="utf-8").splitlines() == [
        "instalacao,unid_leit,equipamento,dta_leit_pr",
        "ATE0000002,001,EQ001,16.04.2026",
    ]


def test_compact_medidor_raw_exports_can_include_iq09_when_requested(tmp_path: Path) -> None:
    raw_dir = tmp_path / "runs" / "run-medidor" / "medidor" / "raw"
    raw_dir.mkdir(parents=True)
    group_map_path = tmp_path / "gruporegsap.xlsx"
    output_path = tmp_path / "compactado.csv"
    _write_xlsx(group_map_path, ["Grp.registrad.", "Tipo"], [["AD30002N", "Digital"], ["DA01010", "Analógico"]])
    (raw_dir / "el31_medidor_20260416_run-medidor_001.txt").write_text(
        "Instalação\tEquipamento\nATE0000002\tEQ001\nMTE0012436\tEQ002\n",
        encoding="utf-8",
    )
    (raw_dir / "iq09_medidor_20260416_run-medidor_001.txt").write_text(
        "Equipamento\tGrpReg.\nEQ001\tAD30002N\nEQ002\tDA01010\n",
        encoding="utf-8",
    )

    result = compact_medidor_raw_exports(
        raw_dir=raw_dir,
        group_map_path=group_map_path,
        output_csv_path=output_path,
        include_iq09=True,
    )

    assert result.deduped_rows_written == 2
    assert result.equipments_without_iq09_group == 0
    assert output_path.read_text(encoding="utf-8").splitlines() == [
        "instalacao,equipamento,grp_reg,tipo",
        "ATE0000002,EQ001,AD30002N,Digital",
        "MTE0012436,EQ002,DA01010,Analógico",
    ]


def test_select_el31_layout_uses_recorded_sap_layout_row(monkeypatch) -> None:  # noqa: ANN001
    class _FakeShell:
        def __init__(self) -> None:
            self.actions: list[tuple[str, str]] = []

        def pressToolbarContextButton(self, item_id: str) -> None:  # noqa: N802
            self.actions.append(("toolbar", item_id))

        def selectContextMenuItem(self, item_id: str) -> None:  # noqa: N802
            self.actions.append(("context", item_id))

    class _FakeGrid:
        def __init__(self) -> None:
            self.firstVisibleRow = 0
            self.current_cell: tuple[int, str] | None = None
            self.selectedRows = ""
            self.clicked = False

        def setCurrentCell(self, row: int, column: str) -> None:  # noqa: N802
            self.current_cell = (row, column)

        def clickCurrentCell(self) -> None:  # noqa: N802
            self.clicked = True

    class _FakeSession:
        def __init__(self, shell: _FakeShell, grid: _FakeGrid) -> None:
            self.controls = {
                "wnd[0]/usr/cntlBCALVC_EVENT2_D100_C1/shellcont/shell": shell,
                "wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell": grid,
            }

        def findById(self, item_id: str):  # noqa: ANN201, N802
            return self.controls[item_id]

    monkeypatch.setattr("sap_gui_export_compat.wait_not_busy", lambda **kwargs: None)
    shell = _FakeShell()
    grid = _FakeGrid()

    MedidorExtractor()._select_el31_layout(
        session=_FakeSession(shell=shell, grid=grid),
        cfg={},
        wait_timeout_seconds=120,
    )

    assert shell.actions == [("toolbar", "&MB_VARIANT"), ("context", "&LOAD")]
    assert grid.firstVisibleRow == 100
    assert grid.current_cell == (105, "TEXT")
    assert grid.selectedRows == "105"
    assert grid.clicked is True


def test_run_el31_with_validation_retries_empty_export(monkeypatch, tmp_path: Path) -> None:  # noqa: ANN001
    extractor = MedidorExtractor()
    attempts: list[int] = []

    def _fake_run_el31(**kwargs):  # noqa: ANN003
        attempts.append(1)
        output_path = kwargs["output_path"]
        if len(attempts) == 1:
            output_path.write_text("Instalação\tEquipamento\n", encoding="utf-8")
            return
        output_path.write_text("Instalação\tEquipamento\nATE0000002\tEQ001\n", encoding="utf-8")

    monkeypatch.setattr(extractor, "_run_el31", _fake_run_el31)
    rows, equipments = extractor._run_el31_with_validation(
        session=object(),
        installations=["ATE0000002"],
        output_path=tmp_path / "el31.txt",
        period_from="01.01.2025",
        period_to="16.04.2026",
        cfg={},
        demandante_cfg={},
        logger=type("Logger", (), {"warning": lambda *args, **kwargs: None})(),
        wait_timeout_seconds=1,
        max_attempts=2,
    )

    assert len(attempts) == 2
    assert equipments == ["EQ001"]
    assert rows == [{"instalacao": "ATE0000002", "unid_leit": "", "equipamento": "EQ001", "dta_leit_pr": ""}]


def test_run_iq09_with_validation_retries_empty_export(monkeypatch, tmp_path: Path) -> None:  # noqa: ANN001
    extractor = MedidorExtractor()
    attempts: list[int] = []

    def _fake_run_iq09(**kwargs):  # noqa: ANN003
        attempts.append(1)
        output_path = kwargs["output_path"]
        if len(attempts) == 1:
            output_path.write_text("Equipamento\tGrpReg.\n", encoding="utf-8")
            return
        output_path.write_text("Equipamento\tGrpReg.\nEQ001\tAD30002N\n", encoding="utf-8")

    monkeypatch.setattr(extractor, "_run_iq09", _fake_run_iq09)
    extractor._run_iq09_with_validation(
        session=object(),
        equipments=["EQ001"],
        output_path=tmp_path / "iq09.txt",
        cfg={},
        demandante_cfg={},
        logger=type("Logger", (), {"warning": lambda *args, **kwargs: None})(),
        wait_timeout_seconds=1,
        max_attempts=2,
    )

    assert len(attempts) == 2


def test_medidor_extractor_runs_el31_then_iq09_and_writes_manifest(monkeypatch, tmp_path: Path) -> None:  # noqa: ANN001
    installations_path = tmp_path / "instalacaosp.xlsx"
    group_map_path = tmp_path / "gruporegsap.xlsx"
    _write_xlsx(installations_path, ["INSTALACAO"], [["ATE0000002"], ["MTE0012436"]])
    _write_xlsx(group_map_path, ["Grp.registrad.", "Tipo"], [["AD30002N", "Digital"], ["DA01010", "Analógico"]])
    calls: list[str] = []

    def _fake_el31(self, **kwargs):  # noqa: ANN001, ANN003
        calls.append(f"el31:{len(kwargs['installations'])}")
        equipment_by_installation = {"ATE0000002": "EQ001", "MTE0012436": "EQ002"}
        lines = ["Instalação\tEquipamento"]
        lines.extend(
            f"{installation}\t{equipment_by_installation[installation]}"
            for installation in kwargs["installations"]
        )
        kwargs["output_path"].write_text("\n".join(lines) + "\n", encoding="utf-8")

    def _fake_iq09(self, **kwargs):  # noqa: ANN001, ANN003
        calls.append(f"iq09:{len(kwargs['equipments'])}")
        kwargs["output_path"].write_text("Equipamento\tGrpReg.\nEQ001\tAD30002N\nEQ002\tDA01010\n", encoding="utf-8")

    monkeypatch.setattr(MedidorExtractor, "_run_el31", _fake_el31)
    monkeypatch.setattr(MedidorExtractor, "_run_iq09", _fake_iq09)

    manifest = MedidorExtractor().execute(
        output_root=tmp_path,
        run_id="run-medidor",
        demandante="MEDIDOR",
        session=object(),
        logger=type("Logger", (), {"info": lambda *args, **kwargs: None})(),
        config={
            "global": {"wait_timeout_seconds": 120},
            "medidor": {
                "el31_chunk_size": 2000,
                "iq09_chunk_size": 5000,
                "demandantes": {
                    "MEDIDOR": {
                        "installations_path": str(installations_path),
                        "group_map_path": str(group_map_path),
                        "el31_chunk_size": 1,
                        "iq09_chunk_size": 5000,
                    }
                },
            },
        },
        installations_path=installations_path,
        group_map_path=group_map_path,
        today=date(2026, 4, 16),
    )

    assert calls == ["el31:1", "el31:1", "iq09:2"]
    assert manifest.status == "success"
    assert manifest.total_installations == 2
    assert manifest.total_equipments == 2
    assert manifest.el31_chunk_size == 1
    assert manifest.el31_chunk_count == 2
    assert manifest.rows_written == 2
    assert len(manifest.raw_el31_paths) == 2
    assert Path(manifest.manifest_path).exists()
    assert Path(manifest.final_csv_path).exists()


def test_medidor_extractor_resumes_from_existing_el31_and_iq09_raws(monkeypatch, tmp_path: Path) -> None:  # noqa: ANN001
    installations_path = tmp_path / "instalacaosp.xlsx"
    group_map_path = tmp_path / "gruporegsap.xlsx"
    _write_xlsx(installations_path, ["INSTALACAO"], [["ATE0000002"], ["MTE0012436"]])
    _write_xlsx(group_map_path, ["Grp.registrad.", "Tipo"], [["AD30002N", "Digital"], ["DA01010", "Analógico"]])
    raw_dir = tmp_path / "runs" / "run-medidor" / "medidor" / "raw"
    raw_dir.mkdir(parents=True)
    (raw_dir / "el31_medidor_20260416_run-medidor_001.txt").write_text(
        "Instalação\tEquipamento\nATE0000002\tEQ001\n",
        encoding="utf-8",
    )
    (raw_dir / "el31_medidor_20260416_run-medidor_002.txt").write_text(
        "Instalação\tEquipamento\nMTE0012436\tEQ002\n",
        encoding="utf-8",
    )
    (raw_dir / "iq09_medidor_20260416_run-medidor_001.txt").write_text(
        "N� s�rie\tStatUsu�r.\tGrpReg.\nEQ001\tINST\tAD30002N\nEQ002\tINST\tDA01010\n",
        encoding="utf-8",
    )

    def _unexpected_sap_call(self, **kwargs):  # noqa: ANN001, ANN003
        raise AssertionError("SAP extraction should not run when valid raw chunks already exist.")

    monkeypatch.setattr(MedidorExtractor, "_run_el31", _unexpected_sap_call)
    monkeypatch.setattr(MedidorExtractor, "_run_iq09", _unexpected_sap_call)

    manifest = MedidorExtractor().execute(
        output_root=tmp_path,
        run_id="run-medidor",
        demandante="MEDIDOR",
        session=object(),
        logger=type("Logger", (), {"info": lambda *args, **kwargs: None, "warning": lambda *args, **kwargs: None})(),
        config={
            "global": {"wait_timeout_seconds": 120},
            "medidor": {
                "el31_chunk_size": 2000,
                "iq09_chunk_size": 5000,
                "demandantes": {
                    "MEDIDOR": {
                        "installations_path": str(installations_path),
                        "group_map_path": str(group_map_path),
                        "el31_chunk_size": 1,
                        "iq09_chunk_size": 2,
                    }
                },
            },
        },
        installations_path=installations_path,
        group_map_path=group_map_path,
        today=date(2026, 4, 17),
    )

    compact_csv_path = tmp_path / "runs" / "run-medidor" / "medidor" / "normalized" / "medidor_raw_compactado.csv"
    assert compact_csv_path.read_text(encoding="utf-8").splitlines() == [
        "instalacao,unid_leit,equipamento,dta_leit_pr",
        "ATE0000002,,EQ001,",
        "MTE0012436,,EQ002,",
    ]
    assert manifest.raw_el31_paths == [
        str(raw_dir / "el31_medidor_20260416_run-medidor_001.txt"),
        str(raw_dir / "el31_medidor_20260416_run-medidor_002.txt"),
    ]
    assert manifest.raw_iq09_paths == [str(raw_dir / "iq09_medidor_20260416_run-medidor_001.txt")]
    assert Path(manifest.final_csv_path).read_text(encoding="utf-8").splitlines() == [
        "instalacao,equipamento,grp_reg,tipo",
        "ATE0000002,EQ001,AD30002N,Digital",
        "MTE0012436,EQ002,DA01010,Analógico",
    ]


def test_run_medidor_sp_forces_sap_logon_pad_rp1(monkeypatch, tmp_path: Path) -> None:  # noqa: ANN001
    captured: dict[str, object] = {}

    monkeypatch.setattr(
        medidor_module,
        "load_export_config",
        lambda path: {
            "global": {
                "logon_pad": {
                    "enabled": False,
                    "connection_description": "wrong",
                },
            },
            "medidor": {"demandantes": {"MEDIDOR": {}}},
        },
    )

    class _Provider:
        def get_session(self, *, config, logger):  # noqa: ANN001
            captured["config"] = config
            return object()

    def _fake_execute(self, **kwargs):  # noqa: ANN001, ANN003
        captured["demandante"] = kwargs["demandante"]
        return medidor_module.MedidorManifest(
            status="success",
            run_id=str(kwargs["run_id"]),
            demandante=str(kwargs["demandante"]),
            reference="20260416",
            period_from="01.01.2025",
            period_to="16.04.2026",
            input_installations_path="",
            group_map_path="",
            manifest_path="manifest.json",
        )

    monkeypatch.setattr(MedidorExtractor, "execute", _fake_execute)

    result = run_medidor_demandante(
        run_id="run-medidor-sp",
        demandante="MEDIDOR_SP",
        output_root=tmp_path,
        config_path=Path("sap_iw69_batch_config.json"),
        session_provider=_Provider(),
    )

    config = captured["config"]
    assert isinstance(config, dict)
    assert config["global"]["logon_pad"]["enabled"] is True
    assert config["global"]["logon_pad"]["connection_description"] == "H181 RP1 ENEL SP CCS Produção (without SSO)"
    assert config["global"]["connection_name"] == "H181 RP1 ENEL SP CCS Produção (without SSO)"
    assert captured["demandante"] == "MEDIDOR_SP"
    assert result.status == "success"
