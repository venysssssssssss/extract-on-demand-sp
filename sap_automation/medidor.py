from __future__ import annotations

import csv
import importlib
import json
import re
import time
import unicodedata
from dataclasses import asdict, dataclass, field
from datetime import date
from pathlib import Path
from typing import Any

from .config import load_export_config
from .runtime_logging import configure_run_logger
from .sap_helpers import (
    resolve_first_existing,
    set_first_existing_text,
    set_text,
    wait_for_file,
)

_MEDIDOR_SP_RP1_CONNECTION = "H181 RP1 ENEL SP CCS Produção (without SSO)"
_MEDIDOR_SP_DEMANDANTES = {"MEDIDOR", "MEDIDOR_SP"}


def _openpyxl_module() -> Any:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to execute MEDIDOR workbook automation. Install dependencies with poetry."
        ) from exc
    return openpyxl


def normalize_table_header(value: str) -> str:
    text = str(value or "")
    if "Ã" in text or "Â" in text:
        try:
            text = text.encode("cp1252").decode("utf-8")
        except UnicodeError:
            pass
    normalized = unicodedata.normalize("NFKD", text)
    ascii_value = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return "_".join(
        part
        for part in "".join(ch.lower() if ch.isalnum() else " " for ch in ascii_value).split()
        if part
    )


def chunk_values(values: list[str], chunk_size: int) -> list[list[str]]:
    if chunk_size <= 0:
        raise ValueError("chunk_size must be greater than zero.")
    return [values[index : index + chunk_size] for index in range(0, len(values), chunk_size)]


def compute_medidor_el31_period(today: date | None = None) -> tuple[str, str, str]:
    current_day = today or date.today()
    start = date(current_day.year - 1, 1, 1)
    return start.strftime("%d.%m.%Y"), current_day.strftime("%d.%m.%Y"), current_day.strftime("%Y%m%d")


def load_workbook_column(path: Path, *, column_name: str, sheet_name: str = "") -> list[str]:
    workbook_path = path.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    openpyxl = _openpyxl_module()
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    header_values = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    normalized_target = normalize_table_header(column_name)
    normalized_headers = [normalize_table_header(str(value or "")) for value in header_values]
    if normalized_target not in normalized_headers:
        raise RuntimeError(
            f"Workbook {workbook_path} is missing required column '{column_name}'. "
            f"Available columns: {header_values}"
        )
    column_index = normalized_headers.index(normalized_target) + 1
    values: list[str] = []
    seen: set[str] = set()
    for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, values_only=True):
        value = str(row[0] or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        values.append(value)
    return values


def load_csv_column(path: Path, *, column_name: str) -> list[str]:
    csv_path = path.expanduser().resolve()
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV not found: {csv_path}")
    normalized_target = normalize_table_header(column_name)
    values: list[str] = []
    seen: set[str] = set()
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise RuntimeError(f"CSV {csv_path} is missing a header row.")
        normalized_headers = [normalize_table_header(name) for name in reader.fieldnames]
        if normalized_target not in normalized_headers:
            raise RuntimeError(
                f"CSV {csv_path} is missing required column '{column_name}'. "
                f"Available columns: {reader.fieldnames}"
            )
        source_name = reader.fieldnames[normalized_headers.index(normalized_target)]
        for row in reader:
            value = str(row.get(source_name, "") or "").strip()
            if value and value not in seen:
                seen.add(value)
                values.append(value)
    return values


def load_grpreg_type_map(path: Path, *, sheet_name: str = "") -> dict[str, str]:
    workbook_path = path.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    openpyxl = _openpyxl_module()
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"Workbook {workbook_path} is empty.")
    headers = [normalize_table_header(str(value or "")) for value in rows[0]]
    grp_candidates = ("grp_registrad", "grp_reg", "grpreg", "grupo_registrador")
    type_candidates = ("tipo", "type")
    grp_index = next((headers.index(item) for item in grp_candidates if item in headers), -1)
    type_index = next((headers.index(item) for item in type_candidates if item in headers), -1)
    if grp_index < 0 or type_index < 0:
        raise RuntimeError(
            f"Workbook {workbook_path} must contain columns 'Grp.registrad.' and 'Tipo'. "
            f"Available columns: {rows[0]}"
        )
    mapping: dict[str, str] = {}
    for row in rows[1:]:
        group_value = str(row[grp_index] or "").strip().upper()
        type_value = str(row[type_index] or "").strip()
        if group_value and group_value not in mapping:
            mapping[group_value] = type_value
    return mapping


_MEDIDOR_TXT_ENCODINGS: tuple[str, ...] = ("cp1252", "utf-8-sig", "utf-8", "latin-1")
_MEDIDOR_TXT_DELIMITERS: tuple[str, ...] = ("\t", "|", ";", ",")


def _canonical_medidor_header(value: str, index: int) -> str:
    key = normalize_table_header(value)
    compact_key = key.replace("_", "")
    if compact_key.startswith("instala") and compact_key.endswith("o"):
        return "instalacao"
    if compact_key in {"unidleit", "unidadeleitura", "unidleitura"}:
        return "unid_leit"
    if compact_key in {
        "equipamento",
        "equipment",
        "equip",
        "eqtl",
        "eqt",
        "eq",
        "nserie",
        "nser",
        "nsrie",
        "noserie",
        "numeroserie",
        "numerodeserie",
        "sernr",
    }:
        return "equipamento"
    if compact_key in {"grpreg", "grpregistrad", "gruporegistrador"}:
        return "grp_reg"
    if compact_key in {"dtaleitpr", "dataleitpr", "dtaleitprev", "dataleitprev"}:
        return "dta_leit_pr"
    return key or f"unnamed_{index}"


def _clean_medidor_txt_row(row: list[str]) -> list[str]:
    cleaned = [str(item or "").strip() for item in row]
    while cleaned and not cleaned[0]:
        cleaned.pop(0)
    while cleaned and not cleaned[-1]:
        cleaned.pop()
    return cleaned


def _is_medidor_separator_row(row: list[str]) -> bool:
    values = [value.strip() for value in row if value.strip()]
    if not values:
        return True
    return all(set(value) <= {"-", "=", "_"} for value in values)


def _read_medidor_delimited_rows(path: Path) -> list[list[str]]:
    errors: list[str] = []
    decoded_without_header: list[str] = []
    for encoding in _MEDIDOR_TXT_ENCODINGS:
        try:
            text = path.read_text(encoding=encoding)
        except Exception as exc:
            errors.append(f"encoding={encoding}: {exc}")
            continue
        lines = [line for line in text.splitlines() if line.strip()]
        if not lines:
            return []
        best_rows: list[list[str]] = []
        best_header_index = -1
        best_score = -1
        for delimiter in _MEDIDOR_TXT_DELIMITERS:
            rows = [_clean_medidor_txt_row(row) for row in csv.reader(lines, delimiter=delimiter)]
            for row_index, row in enumerate(rows):
                if _is_medidor_separator_row(row):
                    continue
                header = [_canonical_medidor_header(item, index) for index, item in enumerate(row)]
                score = int("equipamento" in header) * 10 + int("instalacao" in header)
                if score > best_score:
                    best_rows = rows
                    best_header_index = row_index
                    best_score = score
        whitespace_rows = [
            _clean_medidor_txt_row(re.split(r"\s{2,}|\t+", line.strip()))
            for line in lines
        ]
        for row_index, row in enumerate(whitespace_rows):
            if _is_medidor_separator_row(row):
                continue
            header = [_canonical_medidor_header(item, index) for index, item in enumerate(row)]
            score = int("equipamento" in header) * 10 + int("instalacao" in header)
            if score > best_score:
                best_rows = whitespace_rows
                best_header_index = row_index
                best_score = score
        if best_score >= 10 and best_header_index >= 0:
            return best_rows[best_header_index:]
        decoded_without_header.append(f"encoding={encoding}: decoded but no MEDIDOR header found")
    if decoded_without_header:
        return []
    raise RuntimeError(f"Could not decode MEDIDOR TXT export {path}. " + " | ".join(errors[:4]))


def _read_export_rows(path: Path) -> list[dict[str, str]]:
    rows = _read_medidor_delimited_rows(path)
    if not rows:
        return []
    header = [_canonical_medidor_header(item, index) for index, item in enumerate(rows[0])]
    result: list[dict[str, str]] = []
    for row in rows[1:]:
        if _is_medidor_separator_row(row):
            continue
        payload: dict[str, str] = {}
        for index, key in enumerate(header):
            payload[key] = str(row[index] if index < len(row) else "").strip()
        if any(payload.values()):
            result.append(payload)
    return result


def _first_present(row: dict[str, str], candidates: tuple[str, ...]) -> str:
    for candidate in candidates:
        value = str(row.get(candidate, "") or "").strip()
        if value:
            return value
    return ""


def collect_equipments_from_el31_export(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    rows = _read_export_rows(path)
    normalized_rows: list[dict[str, str]] = []
    equipments: list[str] = []
    seen: set[str] = set()
    for row in rows:
        installation = _first_present(row, ("instalacao", "instalacoes", "installation", "anlage"))
        reading_unit = _first_present(row, ("unid_leit", "unidade_leitura", "unidade_de_leitura"))
        equipment = _first_present(row, ("equipamento", "equipment", "equip", "no_equipamento", "n_equipamento"))
        planned_read_date = _first_present(row, ("dta_leit_pr", "data_leit_pr", "data_leitura_prevista"))
        if not equipment:
            continue
        normalized_rows.append(
            {
                "instalacao": installation,
                "unid_leit": reading_unit,
                "equipamento": equipment,
                "dta_leit_pr": planned_read_date,
            }
        )
        if equipment not in seen:
            seen.add(equipment)
            equipments.append(equipment)
    return normalized_rows, equipments


def collect_iq09_grpreg_by_equipment(paths: list[Path]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for path in paths:
        for row in _read_export_rows(path):
            equipment = _first_present(
                row,
                ("equipamento", "equipment", "equip", "numero_de_serie", "n_serie", "sernr", "no_serie"),
            )
            grp_reg = _first_present(
                row,
                ("grpreg", "grpreg", "grp_reg", "grp_registrad", "grupo_registrador", "grp_registrador"),
            )
            if equipment and grp_reg and equipment not in mapping:
                mapping[equipment] = grp_reg.upper()
    return mapping


def deduplicate_el31_rows_by_equipment(el31_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    deduped_rows: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in el31_rows:
        equipment = str(row.get("equipamento", "")).strip()
        if not equipment or equipment in seen:
            continue
        seen.add(equipment)
        deduped_rows.append(row)
    return deduped_rows


def write_medidor_final_csv(
    *,
    el31_rows: list[dict[str, str]],
    iq09_grpreg_by_equipment: dict[str, str],
    grpreg_type_map: dict[str, str],
    output_path: Path,
) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows_written = 0
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["instalacao", "equipamento", "grp_reg", "tipo"],
        )
        writer.writeheader()
        for row in el31_rows:
            equipment = str(row.get("equipamento", "")).strip()
            grp_reg = iq09_grpreg_by_equipment.get(equipment, "")
            writer.writerow(
                {
                    "instalacao": str(row.get("instalacao", "")).strip(),
                    "equipamento": equipment,
                    "grp_reg": grp_reg,
                    "tipo": grpreg_type_map.get(grp_reg.upper(), ""),
                }
            )
            rows_written += 1
    return rows_written


def write_medidor_el31_compact_csv(
    *,
    el31_rows: list[dict[str, str]],
    output_path: Path,
) -> int:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows_written = 0
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["instalacao", "unid_leit", "equipamento", "dta_leit_pr"])
        writer.writeheader()
        for row in el31_rows:
            writer.writerow(
                {
                    "instalacao": str(row.get("instalacao", "")).strip(),
                    "unid_leit": str(row.get("unid_leit", "")).strip(),
                    "equipamento": str(row.get("equipamento", "")).strip(),
                    "dta_leit_pr": str(row.get("dta_leit_pr", "")).strip(),
                }
            )
            rows_written += 1
    return rows_written


@dataclass(frozen=True)
class MedidorRawCompactResult:
    status: str
    raw_dir: str
    output_csv_path: str
    manifest_path: str
    group_map_path: str
    el31_raw_paths: list[str]
    el31_raw_paths_skipped: list[str]
    iq09_raw_paths: list[str]
    el31_rows_read: int
    deduped_rows_written: int
    duplicate_equipments_removed: int
    equipments_without_iq09_group: int
    equipments_without_type: int

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def discover_medidor_raw_txt_paths(raw_dir: Path) -> tuple[list[Path], list[Path]]:
    resolved_raw_dir = raw_dir.expanduser().resolve()
    if not resolved_raw_dir.exists():
        raise FileNotFoundError(f"MEDIDOR raw directory not found: {resolved_raw_dir}")
    if not resolved_raw_dir.is_dir():
        raise NotADirectoryError(f"MEDIDOR raw path is not a directory: {resolved_raw_dir}")

    txt_paths = sorted(path for path in resolved_raw_dir.iterdir() if path.is_file() and path.suffix.lower() == ".txt")
    el31_paths = [path for path in txt_paths if path.name.lower().startswith("el31_medidor_")]
    iq09_paths = [path for path in txt_paths if path.name.lower().startswith("iq09_medidor_")]
    if not el31_paths:
        raise RuntimeError(f"No EL31 MEDIDOR raw TXT files found in: {resolved_raw_dir}")
    return el31_paths, iq09_paths


def find_existing_medidor_raw_chunk(
    *,
    raw_dir: Path,
    prefix: str,
    run_id: str,
    chunk_index: int,
    preferred_path: Path,
) -> Path | None:
    if preferred_path.exists() and preferred_path.is_file() and preferred_path.stat().st_size > 0:
        return preferred_path
    candidates = [
        path
        for path in raw_dir.glob(f"{prefix}_medidor_*_{run_id}_{chunk_index:03d}.txt")
        if path.is_file() and path.stat().st_size > 0
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def compact_medidor_raw_exports(
    *,
    raw_dir: Path,
    group_map_path: Path | None = None,
    output_csv_path: Path | None = None,
    manifest_path: Path | None = None,
    include_iq09: bool = False,
) -> MedidorRawCompactResult:
    resolved_raw_dir = raw_dir.expanduser().resolve()
    resolved_group_map_path = group_map_path.expanduser().resolve() if group_map_path is not None else None
    el31_paths, iq09_paths = discover_medidor_raw_txt_paths(resolved_raw_dir)
    if include_iq09 and not iq09_paths:
        raise RuntimeError(f"No IQ09 MEDIDOR raw TXT files found in: {resolved_raw_dir}")
    if include_iq09 and resolved_group_map_path is None:
        raise RuntimeError("group_map_path is required when include_iq09=True.")
    grpreg_type_map = load_grpreg_type_map(resolved_group_map_path) if resolved_group_map_path is not None else {}

    deduped_el31_rows: list[dict[str, str]] = []
    seen_equipments: set[str] = set()
    el31_rows_read = 0
    duplicate_equipments_removed = 0
    skipped_el31_paths: list[str] = []
    for path in el31_paths:
        try:
            rows, _ = collect_equipments_from_el31_export(path)
        except RuntimeError as exc:
            skipped_el31_paths.append(f"{path}: {exc}")
            continue
        if not rows:
            skipped_el31_paths.append(f"{path}: no rows with equipment found")
            continue
        el31_rows_read += len(rows)
        for row in rows:
            equipment = str(row.get("equipamento", "")).strip()
            if not equipment:
                continue
            if equipment in seen_equipments:
                duplicate_equipments_removed += 1
                continue
            seen_equipments.add(equipment)
            deduped_el31_rows.append(row)

    if not deduped_el31_rows:
        raise RuntimeError(f"EL31 MEDIDOR raw TXT files did not produce equipment values: {el31_paths}")

    iq09_grpreg_by_equipment = collect_iq09_grpreg_by_equipment(iq09_paths) if include_iq09 else {}
    resolved_output_csv_path = (
        output_csv_path.expanduser().resolve()
        if output_csv_path is not None
        else resolved_raw_dir.parent / "normalized" / "medidor_raw_compactado.csv"
    )
    if include_iq09:
        rows_written = write_medidor_final_csv(
            el31_rows=deduped_el31_rows,
            iq09_grpreg_by_equipment=iq09_grpreg_by_equipment,
            grpreg_type_map=grpreg_type_map,
            output_path=resolved_output_csv_path,
        )
    else:
        rows_written = write_medidor_el31_compact_csv(
            el31_rows=deduped_el31_rows,
            output_path=resolved_output_csv_path,
        )
    equipments_without_iq09_group = (
        sum(
            1
            for row in deduped_el31_rows
            if not iq09_grpreg_by_equipment.get(str(row.get("equipamento", "")).strip(), "")
        )
        if include_iq09
        else 0
    )
    equipments_without_type = 0
    for row in deduped_el31_rows:
        equipment = str(row.get("equipamento", "")).strip()
        grp_reg = iq09_grpreg_by_equipment.get(equipment, "")
        if grp_reg and not grpreg_type_map.get(grp_reg.upper(), ""):
            equipments_without_type += 1

    resolved_manifest_path = (
        manifest_path.expanduser().resolve()
        if manifest_path is not None
        else resolved_output_csv_path.with_suffix(".manifest.json")
    )
    result = MedidorRawCompactResult(
        status="success",
        raw_dir=str(resolved_raw_dir),
        output_csv_path=str(resolved_output_csv_path),
        manifest_path=str(resolved_manifest_path),
        group_map_path=str(resolved_group_map_path or ""),
        el31_raw_paths=[str(path) for path in el31_paths],
        el31_raw_paths_skipped=skipped_el31_paths,
        iq09_raw_paths=[str(path) for path in iq09_paths],
        el31_rows_read=el31_rows_read,
        deduped_rows_written=rows_written,
        duplicate_equipments_removed=duplicate_equipments_removed,
        equipments_without_iq09_group=equipments_without_iq09_group,
        equipments_without_type=equipments_without_type,
    )
    resolved_manifest_path.parent.mkdir(parents=True, exist_ok=True)
    resolved_manifest_path.write_text(json.dumps(result.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    return result


@dataclass(frozen=True)
class MedidorManifest:
    status: str
    run_id: str
    demandante: str
    reference: str
    period_from: str
    period_to: str
    input_installations_path: str
    group_map_path: str
    total_installations: int = 0
    total_equipments: int = 0
    el31_chunk_size: int = 0
    el31_chunk_count: int = 0
    iq09_chunk_size: int = 0
    iq09_chunk_count: int = 0
    rows_written: int = 0
    raw_el31_path: str = ""
    raw_el31_paths: list[str] = field(default_factory=list)
    raw_iq09_paths: list[str] = field(default_factory=list)
    final_csv_path: str = ""
    manifest_path: str = ""
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


class MedidorExtractor:
    def execute(
        self,
        *,
        output_root: Path,
        run_id: str,
        demandante: str,
        session: Any,
        logger: Any,
        config: dict[str, Any],
        installations_path: Path | None = None,
        installations: list[str] | None = None,
        group_map_path: Path | None = None,
        today: date | None = None,
    ) -> MedidorManifest:
        medidor_cfg = config.get("medidor", {})
        if not isinstance(medidor_cfg, dict):
            raise RuntimeError("Missing medidor section in config JSON.")
        demandante_name = str(demandante or "").strip().upper() or "MEDIDOR"
        demandante_cfg = (
            medidor_cfg.get("demandantes", {}).get(demandante_name, {})
            if isinstance(medidor_cfg.get("demandantes", {}), dict)
            else {}
        )
        resolved_installations_path = (
            installations_path
            or Path(str(demandante_cfg.get("installations_path", "instalacaosp.xlsx")))
        ).expanduser().resolve()
        resolved_group_map_path = (
            group_map_path
            or Path(str(demandante_cfg.get("group_map_path", "gruporegsap.xlsx")))
        ).expanduser().resolve()
        installation_column = str(demandante_cfg.get("installation_column", "INSTALACAO"))
        if installations is None:
            if resolved_installations_path.suffix.lower() == ".csv":
                installations = load_csv_column(resolved_installations_path, column_name=installation_column)
            else:
                installations = load_workbook_column(
                    resolved_installations_path,
                    column_name=installation_column,
                    sheet_name=str(demandante_cfg.get("installations_sheet_name", "")).strip(),
                )
        else:
            installations = _deduplicate_medidor_tokens(installations)
        grpreg_type_map = load_grpreg_type_map(
            resolved_group_map_path,
            sheet_name=str(demandante_cfg.get("group_map_sheet_name", "")).strip(),
        )
        period_from, period_to, reference = compute_medidor_el31_period(today=today)
        run_root = output_root.expanduser().resolve() / "runs" / run_id / "medidor"
        raw_dir = run_root / "raw"
        normalized_dir = run_root / "normalized"
        metadata_dir = run_root / "metadata"
        raw_dir.mkdir(parents=True, exist_ok=True)
        normalized_dir.mkdir(parents=True, exist_ok=True)
        metadata_dir.mkdir(parents=True, exist_ok=True)

        wait_timeout_seconds = float(config.get("global", {}).get("wait_timeout_seconds", 120.0))
        export_retries = int(demandante_cfg.get("export_retries", medidor_cfg.get("export_retries", 2)))
        el31_chunk_size = int(demandante_cfg.get("el31_chunk_size", medidor_cfg.get("el31_chunk_size", 2000)))
        logger.info(
            "Starting MEDIDOR extraction run_id=%s installations=%s el31_chunk_size=%s period_from=%s period_to=%s input=%s group_map=%s",
            run_id,
            len(installations),
            el31_chunk_size,
            period_from,
            period_to,
            resolved_installations_path,
            resolved_group_map_path,
        )
        el31_paths: list[Path] = []
        equipments: list[str] = []
        seen_equipments: set[str] = set()
        deduped_el31_rows: list[dict[str, str]] = []
        installation_chunks = chunk_values(installations, el31_chunk_size)
        for chunk_index, installation_chunk in enumerate(installation_chunks, start=1):
            el31_raw_path = raw_dir / f"el31_medidor_{reference}_{run_id}_{chunk_index:03d}.txt"
            existing_el31_path = find_existing_medidor_raw_chunk(
                raw_dir=raw_dir,
                prefix="el31",
                run_id=run_id,
                chunk_index=chunk_index,
                preferred_path=el31_raw_path,
            )
            logger.info(
                "Running MEDIDOR EL31 chunk index=%s/%s installations=%s output=%s",
                chunk_index,
                len(installation_chunks),
                len(installation_chunk),
                el31_raw_path,
            )
            if existing_el31_path is not None:
                chunk_rows, chunk_equipments = collect_equipments_from_el31_export(existing_el31_path)
                if chunk_equipments:
                    logger.info(
                        "Skipping MEDIDOR EL31 chunk because valid raw export already exists index=%s path=%s equipments=%s",
                        chunk_index,
                        existing_el31_path,
                        len(chunk_equipments),
                    )
                    el31_paths.append(existing_el31_path)
                else:
                    logger.warning(
                        "Existing MEDIDOR EL31 raw export is invalid; re-extracting index=%s path=%s",
                        chunk_index,
                        existing_el31_path,
                    )
                    chunk_rows, chunk_equipments = self._run_el31_with_validation(
                        session=session,
                        installations=installation_chunk,
                        output_path=el31_raw_path,
                        period_from=period_from,
                        period_to=period_to,
                        cfg=medidor_cfg,
                        demandante_cfg=demandante_cfg,
                        logger=logger,
                        wait_timeout_seconds=wait_timeout_seconds,
                        max_attempts=export_retries + 1,
                    )
                    el31_paths.append(el31_raw_path)
            else:
                chunk_rows, chunk_equipments = self._run_el31_with_validation(
                    session=session,
                    installations=installation_chunk,
                    output_path=el31_raw_path,
                    period_from=period_from,
                    period_to=period_to,
                    cfg=medidor_cfg,
                    demandante_cfg=demandante_cfg,
                    logger=logger,
                    wait_timeout_seconds=wait_timeout_seconds,
                    max_attempts=export_retries + 1,
                )
                el31_paths.append(el31_raw_path)
            for row in chunk_rows:
                equipment = str(row.get("equipamento", "")).strip()
                if not equipment or equipment in seen_equipments:
                    continue
                seen_equipments.add(equipment)
                equipments.append(equipment)
                deduped_el31_rows.append(row)
        if not equipments:
            raise RuntimeError(f"EL31 exports did not produce any equipment values: {el31_paths}")

        el31_compact_csv_path = normalized_dir / "medidor_raw_compactado.csv"
        if el31_compact_csv_path.exists():
            logger.info("MEDIDOR EL31 compact CSV already exists path=%s", el31_compact_csv_path)
        else:
            el31_compact_rows = write_medidor_el31_compact_csv(
                el31_rows=deduped_el31_rows,
                output_path=el31_compact_csv_path,
            )
            logger.info(
                "MEDIDOR EL31 compact CSV created path=%s rows=%s",
                el31_compact_csv_path,
                el31_compact_rows,
            )

        iq09_chunk_size = int(demandante_cfg.get("iq09_chunk_size", medidor_cfg.get("iq09_chunk_size", 5000)))
        iq09_paths: list[Path] = []
        for chunk_index, equipment_chunk in enumerate(chunk_values(equipments, iq09_chunk_size), start=1):
            iq09_path = raw_dir / f"iq09_medidor_{reference}_{run_id}_{chunk_index:03d}.txt"
            existing_iq09_path = find_existing_medidor_raw_chunk(
                raw_dir=raw_dir,
                prefix="iq09",
                run_id=run_id,
                chunk_index=chunk_index,
                preferred_path=iq09_path,
            )
            logger.info(
                "Running MEDIDOR IQ09 chunk index=%s equipments=%s output=%s",
                chunk_index,
                len(equipment_chunk),
                iq09_path,
            )
            if existing_iq09_path is not None and collect_iq09_grpreg_by_equipment([existing_iq09_path]):
                logger.info(
                    "Skipping MEDIDOR IQ09 chunk because valid raw export already exists index=%s path=%s",
                    chunk_index,
                    existing_iq09_path,
                )
                iq09_paths.append(existing_iq09_path)
                continue
            if existing_iq09_path is not None:
                logger.warning(
                    "Existing MEDIDOR IQ09 raw export is invalid; re-extracting index=%s path=%s",
                    chunk_index,
                    existing_iq09_path,
                )
            self._run_iq09_with_validation(
                session=session,
                equipments=equipment_chunk,
                output_path=iq09_path,
                cfg=medidor_cfg,
                demandante_cfg=demandante_cfg,
                logger=logger,
                wait_timeout_seconds=wait_timeout_seconds,
                max_attempts=export_retries + 1,
            )
            iq09_paths.append(iq09_path)

        iq09_grpreg_by_equipment = collect_iq09_grpreg_by_equipment(iq09_paths)
        final_csv_path = normalized_dir / f"medidor_{reference}_{run_id}.csv"
        rows_written = write_medidor_final_csv(
            el31_rows=deduped_el31_rows,
            iq09_grpreg_by_equipment=iq09_grpreg_by_equipment,
            grpreg_type_map=grpreg_type_map,
            output_path=final_csv_path,
        )
        manifest_path = metadata_dir / f"medidor_{reference}_{run_id}.manifest.json"
        manifest = MedidorManifest(
            status="success",
            run_id=run_id,
            demandante=demandante_name,
            reference=reference,
            period_from=period_from,
            period_to=period_to,
            input_installations_path=str(resolved_installations_path),
            group_map_path=str(resolved_group_map_path),
            total_installations=len(installations),
            total_equipments=len(equipments),
            el31_chunk_size=el31_chunk_size,
            el31_chunk_count=len(el31_paths),
            iq09_chunk_size=iq09_chunk_size,
            iq09_chunk_count=len(iq09_paths),
            rows_written=rows_written,
            raw_el31_path=str(el31_paths[0]) if el31_paths else "",
            raw_el31_paths=[str(path) for path in el31_paths],
            raw_iq09_paths=[str(path) for path in iq09_paths],
            final_csv_path=str(final_csv_path),
            manifest_path=str(manifest_path),
        )
        manifest_path.write_text(json.dumps(manifest.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
        logger.info(
            "MEDIDOR extraction finished run_id=%s equipments=%s el31_chunks=%s iq09_chunks=%s rows_written=%s final_csv=%s",
            run_id,
            len(equipments),
            len(el31_paths),
            len(iq09_paths),
            rows_written,
            final_csv_path,
        )
        return manifest

    def _run_el31_with_validation(
        self,
        *,
        session: Any,
        installations: list[str],
        output_path: Path,
        period_from: str,
        period_to: str,
        cfg: dict[str, Any],
        demandante_cfg: dict[str, Any],
        logger: Any,
        wait_timeout_seconds: float,
        max_attempts: int,
    ) -> tuple[list[dict[str, str]], list[str]]:
        last_error: Exception | None = None
        for attempt in range(1, max(1, max_attempts) + 1):
            try:
                self._run_el31(
                    session=session,
                    installations=installations,
                    output_path=output_path,
                    period_from=period_from,
                    period_to=period_to,
                    cfg=cfg,
                    demandante_cfg=demandante_cfg,
                    logger=logger,
                    wait_timeout_seconds=wait_timeout_seconds,
                )
                rows, equipments = collect_equipments_from_el31_export(output_path)
                if not equipments:
                    raise RuntimeError(f"EL31 export produced no equipment values: {output_path}")
                return rows, equipments
            except Exception as exc:
                last_error = exc
                logger.warning(
                    "MEDIDOR EL31 chunk attempt failed attempt=%s/%s output=%s error=%s",
                    attempt,
                    max_attempts,
                    output_path,
                    exc,
                )
                if attempt < max_attempts:
                    self._settle_after_failed_attempt(session=session, wait_timeout_seconds=wait_timeout_seconds)
        raise RuntimeError(f"MEDIDOR EL31 export failed after {max_attempts} attempts: {output_path}") from last_error

    def _run_iq09_with_validation(
        self,
        *,
        session: Any,
        equipments: list[str],
        output_path: Path,
        cfg: dict[str, Any],
        demandante_cfg: dict[str, Any],
        logger: Any,
        wait_timeout_seconds: float,
        max_attempts: int,
    ) -> None:
        last_error: Exception | None = None
        for attempt in range(1, max(1, max_attempts) + 1):
            try:
                self._run_iq09(
                    session=session,
                    equipments=equipments,
                    output_path=output_path,
                    cfg=cfg,
                    demandante_cfg=demandante_cfg,
                    logger=logger,
                    wait_timeout_seconds=wait_timeout_seconds,
                )
                mapping = collect_iq09_grpreg_by_equipment([output_path])
                if not mapping:
                    raise RuntimeError(f"IQ09 export produced no equipment/group mappings: {output_path}")
                return
            except Exception as exc:
                last_error = exc
                logger.warning(
                    "MEDIDOR IQ09 chunk attempt failed attempt=%s/%s output=%s error=%s",
                    attempt,
                    max_attempts,
                    output_path,
                    exc,
                )
                if attempt < max_attempts:
                    self._settle_after_failed_attempt(session=session, wait_timeout_seconds=wait_timeout_seconds)
        raise RuntimeError(f"MEDIDOR IQ09 export failed after {max_attempts} attempts: {output_path}") from last_error

    def _run_el31(
        self,
        *,
        session: Any,
        installations: list[str],
        output_path: Path,
        period_from: str,
        period_to: str,
        cfg: dict[str, Any],
        demandante_cfg: dict[str, Any],
        logger: Any,
        wait_timeout_seconds: float,
    ) -> None:
        compat = importlib.import_module("sap_gui_export_compat")
        self._enter_transaction(session, str(cfg.get("el31_transaction_code", "EL31")), wait_timeout_seconds)
        resolve_first_existing(session, ["wnd[0]/usr/btn%PS05003_1000"]).press()
        resolve_first_existing(session, ["wnd[0]/usr/btn%_SEL_INS_%_APP_%-VALU_PUSH"]).press()
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        self._fill_multiselect(
            session=session,
            values=installations,
            logger=logger,
            wait_timeout_seconds=wait_timeout_seconds,
            label="installations",
        )
        resolve_first_existing(session, ["wnd[0]/usr/btn%PS33067_1000"]).press()
        set_text(session, "wnd[0]/usr/ctxtSEL_RDA-LOW", period_from)
        set_text(session, "wnd[0]/usr/ctxtSEL_RDA-HIGH", period_to)
        resolve_first_existing(session, ["wnd[0]/tbar[1]/btn[8]"]).press()
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        self._select_el31_layout(
            session=session,
            cfg=demandante_cfg,
            wait_timeout_seconds=wait_timeout_seconds,
        )
        self._export_text_file(
            session=session,
            output_path=output_path,
            logger=logger,
            wait_timeout_seconds=wait_timeout_seconds,
            shell_ids=[
                "wnd[0]/usr/cntlBCALVC_EVENT2_D100_C1/shellcont/shell",
                "wnd[0]/usr/cntlGRID1/shellcont/shell",
                "wnd[0]/usr/cntlCONTAINER/shellcont/shell",
            ],
        )

    def _run_iq09(
        self,
        *,
        session: Any,
        equipments: list[str],
        output_path: Path,
        cfg: dict[str, Any],
        demandante_cfg: dict[str, Any],
        logger: Any,
        wait_timeout_seconds: float,
    ) -> None:
        compat = importlib.import_module("sap_gui_export_compat")
        self._enter_transaction(session, str(cfg.get("iq09_transaction_code", "IQ09")), wait_timeout_seconds)
        set_text(session, "wnd[0]/usr/ctxtDATUV", "")
        set_text(session, "wnd[0]/usr/ctxtDATUB", "")
        resolve_first_existing(session, ["wnd[0]/usr/btn%_SERNR_%_APP_%-VALU_PUSH"]).press()
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        self._fill_multiselect(
            session=session,
            values=equipments,
            logger=logger,
            wait_timeout_seconds=wait_timeout_seconds,
            label="equipments",
        )
        set_text(session, "wnd[0]/usr/ctxtDATUV", "")
        set_text(session, "wnd[0]/usr/ctxtDATUB", "")
        resolve_first_existing(session, ["wnd[0]/tbar[1]/btn[8]"]).press()
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        self._select_iq09_layout(
            session=session,
            cfg=demandante_cfg,
            wait_timeout_seconds=wait_timeout_seconds,
        )
        self._export_text_file(
            session=session,
            output_path=output_path,
            logger=logger,
            wait_timeout_seconds=wait_timeout_seconds,
            menu_ids=["wnd[0]/mbar/menu[0]/menu[10]/menu[2]", "wnd[0]/mbar/menu[0]/menu[11]/menu[2]"],
        )

    def _enter_transaction(self, session: Any, transaction_code: str, wait_timeout_seconds: float) -> None:
        compat = importlib.import_module("sap_gui_export_compat")
        window = session.findById("wnd[0]")
        try:
            window.resizeWorkingPane(89, 18, False)
        except Exception:
            try:
                window.maximize()
            except Exception:
                pass
        okcd = session.findById("wnd[0]/tbar[0]/okcd")
        okcd.text = f"/n{transaction_code.strip().upper()}"
        window.sendVKey(0)
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)

    @staticmethod
    def _resolve_with_retry(
        *,
        session: Any,
        ids: list[str],
        wait_timeout_seconds: float,
    ) -> Any:
        deadline = time.time() + min(max(wait_timeout_seconds, 1.0), 15.0)
        last_error: Exception | None = None
        while time.time() <= deadline:
            try:
                return resolve_first_existing(session, ids)
            except Exception as exc:
                last_error = exc
                time.sleep(0.25)
        raise RuntimeError(f"Could not resolve SAP element after retry: {ids}") from last_error

    @staticmethod
    def _settle_after_failed_attempt(*, session: Any, wait_timeout_seconds: float) -> None:
        compat = importlib.import_module("sap_gui_export_compat")
        try:
            compat.wait_not_busy(session=session, timeout_seconds=min(wait_timeout_seconds, 10.0))
        except Exception:
            time.sleep(1.0)

    @staticmethod
    def _remove_existing_export_file(output_path: Path) -> None:
        if not output_path.exists():
            return
        try:
            output_path.unlink()
        except OSError as exc:
            raise RuntimeError(f"Could not remove stale export file before SAP export: {output_path}") from exc

    @staticmethod
    def _wait_for_stable_export_file(output_path: Path, *, timeout_seconds: float) -> None:
        deadline = time.time() + max(1.0, timeout_seconds)
        previous_size = -1
        stable_seen_at = 0.0
        while time.time() <= deadline:
            try:
                current_size = output_path.stat().st_size
            except OSError:
                time.sleep(0.2)
                continue
            if current_size <= 0:
                previous_size = current_size
                stable_seen_at = 0.0
                time.sleep(0.2)
                continue
            if current_size == previous_size:
                if stable_seen_at and time.time() - stable_seen_at >= 0.5:
                    return
                if not stable_seen_at:
                    stable_seen_at = time.time()
            else:
                previous_size = current_size
                stable_seen_at = time.time()
            time.sleep(0.2)
        raise RuntimeError(f"Export file was not stable in time: {output_path}")

    def _fill_multiselect(
        self,
        *,
        session: Any,
        values: list[str],
        logger: Any,
        wait_timeout_seconds: float,
        label: str,
    ) -> None:
        compat = importlib.import_module("sap_gui_export_compat")
        resolve_first_existing(session, ["wnd[1]/tbar[0]/btn[16]"]).press()
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        self._copy_values_to_clipboard(values, logger=logger, label=label)
        resolve_first_existing(session, ["wnd[1]/tbar[0]/btn[24]"]).press()
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        resolve_first_existing(session, ["wnd[1]/tbar[0]/btn[0]"]).press()
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        resolve_first_existing(session, ["wnd[1]/tbar[0]/btn[8]"]).press()
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)

    @staticmethod
    def _copy_values_to_clipboard(values: list[str], *, logger: Any, label: str) -> None:
        payload = "\r\n".join(str(value).strip() for value in values if str(value).strip())
        try:
            import win32clipboard  # type: ignore

            win32clipboard.OpenClipboard()
            try:
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardText(payload, win32clipboard.CF_UNICODETEXT)
            finally:
                win32clipboard.CloseClipboard()
        except Exception as exc:
            raise RuntimeError(f"Could not copy {label} to Windows clipboard for SAP multiselect: {exc}") from exc
        logger.info("MEDIDOR clipboard ready label=%s count=%s", label, len(values))

    def _select_el31_layout(self, *, session: Any, cfg: dict[str, Any], wait_timeout_seconds: float) -> None:
        compat = importlib.import_module("sap_gui_export_compat")
        shell = self._resolve_with_retry(
            session=session,
            ids=[
                "wnd[0]/usr/cntlBCALVC_EVENT2_D100_C1/shellcont/shell",
                "wnd[0]/usr/cntlGRID1/shellcont/shell",
                "wnd[0]/usr/cntlCONTAINER/shellcont/shell",
            ],
            wait_timeout_seconds=wait_timeout_seconds,
        )
        shell.pressToolbarContextButton("&MB_VARIANT")
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        shell.selectContextMenuItem("&LOAD")
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        grid = self._resolve_with_retry(
            session=session,
            ids=[
                "wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell",
                "wnd[1]/usr/cntlGRID/shellcont/shell",
            ],
            wait_timeout_seconds=wait_timeout_seconds,
        )
        self._select_layout_grid_row(
            grid=grid,
            row=int(cfg.get("el31_layout_row", 105)),
            column=str(cfg.get("el31_layout_column", "TEXT")),
            first_visible_row=int(cfg.get("el31_layout_first_visible_row", 100)),
            double_click=False,
        )

    def _select_iq09_layout(self, *, session: Any, cfg: dict[str, Any], wait_timeout_seconds: float) -> None:
        compat = importlib.import_module("sap_gui_export_compat")
        self._resolve_with_retry(
            session=session,
            ids=["wnd[0]/mbar/menu[5]/menu[2]/menu[1]"],
            wait_timeout_seconds=wait_timeout_seconds,
        ).select()
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        grid = self._resolve_with_retry(
            session=session,
            ids=[
                "wnd[1]/usr/cntlGRID/shellcont/shell",
                "wnd[1]/usr/ssubD0500_SUBSCREEN:SAPLSLVC_DIALOG:0501/cntlG51_CONTAINER/shellcont/shell",
            ],
            wait_timeout_seconds=wait_timeout_seconds,
        )
        self._select_layout_grid_row(
            grid=grid,
            row=int(cfg.get("iq09_layout_row", 166)),
            column=str(cfg.get("iq09_layout_column", "TEXT")),
            first_visible_row=int(cfg.get("iq09_layout_first_visible_row", 161)),
            double_click=True,
        )

    @staticmethod
    def _select_layout_grid_row(
        *,
        grid: Any,
        row: int,
        column: str,
        first_visible_row: int,
        double_click: bool,
    ) -> None:
        try:
            grid.firstVisibleRow = first_visible_row
        except Exception:
            pass
        grid.setCurrentCell(row, column)
        grid.selectedRows = str(row)
        if double_click:
            grid.doubleClickCurrentCell()
        else:
            grid.clickCurrentCell()

    def _export_text_file(
        self,
        *,
        session: Any,
        output_path: Path,
        logger: Any,
        wait_timeout_seconds: float,
        menu_ids: list[str] | None = None,
        shell_ids: list[str] | None = None,
    ) -> None:
        compat = importlib.import_module("sap_gui_export_compat")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._remove_existing_export_file(output_path)
        if menu_ids:
            try:
                self._resolve_with_retry(
                    session=session,
                    ids=menu_ids,
                    wait_timeout_seconds=wait_timeout_seconds,
                ).select()
                compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
            except Exception:
                if not shell_ids:
                    raise
                self._open_shell_export(session=session, shell_ids=shell_ids, wait_timeout_seconds=wait_timeout_seconds)
        elif shell_ids:
            self._open_shell_export(session=session, shell_ids=shell_ids, wait_timeout_seconds=wait_timeout_seconds)
        else:
            raise RuntimeError("MEDIDOR export requires menu_ids or shell_ids.")

        local_file_radio = self._resolve_with_retry(
            session=session,
            ids=["wnd[1]/usr/subSUBSCREEN_STEPLOOP:SAPLSPO5:0150/sub:SAPLSPO5:0150/radSPOPLI-SELFLAG[1,0]"],
            wait_timeout_seconds=wait_timeout_seconds,
        )
        local_file_radio.select()
        local_file_radio.setFocus()
        self._resolve_with_retry(
            session=session,
            ids=["wnd[1]/tbar[0]/btn[0]", "wnd[1]/tbar[0]/btn[11]"],
            wait_timeout_seconds=wait_timeout_seconds,
        ).press()
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        set_first_existing_text(
            session,
            ["wnd[1]/usr/ctxtDY_PATH", "wnd[2]/usr/ctxtDY_PATH", "wnd[1]/usr/txtDY_PATH", "wnd[2]/usr/txtDY_PATH"],
            str(output_path.parent),
        )
        set_first_existing_text(
            session,
            [
                "wnd[1]/usr/ctxtDY_FILENAME",
                "wnd[2]/usr/ctxtDY_FILENAME",
                "wnd[1]/usr/txtDY_FILENAME",
                "wnd[2]/usr/txtDY_FILENAME",
            ],
            output_path.name,
        )
        self._resolve_with_retry(
            session=session,
            ids=["wnd[1]/tbar[0]/btn[11]", "wnd[1]/tbar[0]/btn[0]", "wnd[2]/tbar[0]/btn[11]", "wnd[2]/tbar[0]/btn[0]"],
            wait_timeout_seconds=wait_timeout_seconds,
        ).press()
        wait_for_file(output_path, timeout_seconds=wait_timeout_seconds)
        self._wait_for_stable_export_file(output_path, timeout_seconds=wait_timeout_seconds)
        logger.info("MEDIDOR export saved output=%s", output_path)

    def _open_shell_export(self, *, session: Any, shell_ids: list[str], wait_timeout_seconds: float) -> None:
        compat = importlib.import_module("sap_gui_export_compat")
        shell = self._resolve_with_retry(
            session=session,
            ids=shell_ids,
            wait_timeout_seconds=wait_timeout_seconds,
        )
        shell.pressToolbarContextButton("&MB_EXPORT")
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)
        shell.selectContextMenuItem("&PC")
        compat.wait_not_busy(session=session, timeout_seconds=wait_timeout_seconds)


def run_medidor_demandante(
    *,
    run_id: str,
    demandante: str,
    config_path: Path,
    output_root: Path,
    installations_path: Path | None = None,
    installations: list[str] | None = None,
    group_map_path: Path | None = None,
    session_provider: Any | None = None,
) -> MedidorManifest:
    config = load_export_config(config_path)
    resolved_demandante = str(demandante or "").strip().upper() or "MEDIDOR"
    if resolved_demandante in _MEDIDOR_SP_DEMANDANTES:
        _ensure_medidor_sp_logon_config(config)
    resolved_output_root = output_root.expanduser().resolve()
    logger, _ = configure_run_logger(output_root=resolved_output_root, run_id=run_id)
    if session_provider is None:
        from .service import create_session_provider

        session_provider = create_session_provider(config)
    session = session_provider.get_session(config=config, logger=logger)
    return MedidorExtractor().execute(
        output_root=resolved_output_root,
        run_id=run_id,
        demandante=resolved_demandante,
        session=session,
        logger=logger,
        config=config,
        installations_path=installations_path,
        installations=installations,
        group_map_path=group_map_path,
    )


def _ensure_medidor_sp_logon_config(config: dict[str, Any]) -> None:
    global_cfg = config.setdefault("global", {})
    if not isinstance(global_cfg, dict):
        raise RuntimeError("Config global section must be a JSON object.")
    logon_pad_cfg = global_cfg.setdefault("logon_pad", {})
    if not isinstance(logon_pad_cfg, dict):
        raise RuntimeError("Config global.logon_pad section must be a JSON object.")
    logon_pad_cfg["enabled"] = True
    logon_pad_cfg.setdefault("workspace_name", "00 SAP ERP")
    logon_pad_cfg["connection_description"] = _MEDIDOR_SP_RP1_CONNECTION
    logon_pad_cfg.setdefault("multiple_logon_action", "continue")
    logon_pad_cfg.setdefault("ui_fallback_enabled", True)
    global_cfg["connection_name"] = _MEDIDOR_SP_RP1_CONNECTION
    global_cfg.setdefault("auto_login", True)


def _deduplicate_medidor_tokens(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        token = str(value or "").strip()
        if token and token not in seen:
            seen.add(token)
            result.append(token)
    return result
