from __future__ import annotations

import argparse
from contextlib import nullcontext
import csv
import json
import multiprocessing
import os
import queue
import re
import shutil
import threading
import time
import unicodedata
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

from .config import load_export_config
from .runtime_logging import configure_run_logger
from .sap_helpers import resolve_first_existing_with_id, set_selected, set_text, wait_not_busy as _sap_wait_not_busy

_IW51_TRANSACTION_CODE = "IW51"
_IW51_FIXED_QMART = "CA"
_IW51_FIXED_QWRNUM = "389496787"
_IW51_OKCODE_ID = "wnd[0]/tbar[0]/okcd"
_IW51_MAIN_WINDOW_ID = "wnd[0]"
_IW51_QMART_ID = "wnd[0]/usr/ctxtRIWO00-QMART"
_IW51_QWRNUM_ID = "wnd[0]/usr/ctxtRIWO00-QWRNUM"
_IW51_TAB_19_ID = r"wnd[0]/usr/tabsTAB_GROUP_10/tabp10\TAB19"
_IW51_TAB_01_ID = r"wnd[0]/usr/tabsTAB_GROUP_10/tabp10\TAB01"
_IW51_TPLNR_ID = (
    r"wnd[0]/usr/tabsTAB_GROUP_10/tabp10\TAB19/ssubSUB_GROUP_10:SAPLIQS0:7235/"
    r"subCUSTOM_SCREEN:SAPLIQS0:7212/subSUBSCREEN_1:SAPLIQS0:7322/subOBJEKT:SAPLIWO1:0100/"
    r"ctxtRIWO1-TPLNR"
)
_IW51_INSTALACAO_ID = (
    r"wnd[0]/usr/tabsTAB_GROUP_10/tabp10\TAB19/ssubSUB_GROUP_10:SAPLIQS0:7235/"
    r"subCUSTOM_SCREEN:SAPLIQS0:7212/subSUBSCREEN_3:SAPLIQS0:7900/subUSER0001:SAPLXQQM:0114/"
    r"ctxtQMEL-ZZANLAGE"
)
_IW51_PN_ID = (
    r"wnd[0]/usr/tabsTAB_GROUP_10/tabp10\TAB01/ssubSUB_GROUP_10:SAPLIQS0:7235/"
    r"subCUSTOM_SCREEN:SAPLIQS0:7212/subSUBSCREEN_1:SAPLIQS0:7515/subANSPRECH:SAPLIPAR:0700/"
    r"tabsTSTRIP_700/tabpKUND/ssubTSTRIP_SCREEN:SAPLIPAR:0130/subADRESSE:SAPLIPAR:0150/"
    r"ctxtVIQMEL-KUNUM"
)
_IW51_TIPOLOGIA_ID = (
    r"wnd[0]/usr/subSCREEN_1:SAPLIQS0:1060/subNOTIF_TYPE:SAPLIQS0:1061/txtVIQMEL-QMTXT"
)
_IW51_STATUS_BUTTON_ID = "wnd[0]/usr/subSCREEN_1:SAPLIQS0:1060/btnANWENDERSTATUS"
_IW51_STATUS_FIRST_ID = "wnd[1]/usr/tblSAPLBSVATC_E/radJ_STMAINT-ANWS[0,1]"
_IW51_STATUS_SECOND_ID = "wnd[1]/usr/tblSAPLBSVATC_E/radJ_STMAINT-ANWS[0,4]"
_IW51_STATUS_CONFIRM_ID = "wnd[1]/tbar[0]/btn[0]"
_IW51_SAVE_BUTTON_ID = "wnd[0]/tbar[1]/btn[16]"
_IW51_SAVE_CONFIRM_ID = "wnd[1]/tbar[0]/btn[0]"
_IW51_REQUIRED_HEADERS = ("pn", "instalacao", "tipologia")
_IW51_DONE_HEADER = "FEITO"
_IW51_DONE_VALUE = "SIM"
_IW51_ENTRY_FIELD_IDS = [_IW51_QMART_ID, _IW51_QWRNUM_ID]
_IW51_FAST_FAIL_THRESHOLD_SECONDS = 0.5
_IW51_PROGRESS_LOG_INTERVAL = 50
_IW51_SAP_COM_LOCK = threading.RLock()
_IW51_EXECUTION_MODES = {"sequential", "interleaved", "true_parallel", "process_parallel"}
_IW51_HEADERLESS_DEFAULT_COLUMNS = {"pn": 1, "instalacao": 2, "tipologia": 3}
_IW51_HEADERLESS_DONE_COLUMN = 4
_IW51_LEDGER_FIELDNAMES = [
    "row_index",
    "worker_index",
    "pn",
    "instalacao",
    "tipologia",
    "status",
    "attempt",
    "elapsed_seconds",
    "error",
]
_IW51_LOG_COMPLETED_RE = re.compile(
    r"IW51 item completed worker=(?P<worker>\d+) row=(?P<row>\d+) "
    r"pn=(?P<pn>.*?) instalacao=(?P<instalacao>.*?) tipologia=(?P<tipologia>.*)$"
)
_IW51_LOG_OK_RE = re.compile(
    r"IW51 item OK worker=(?P<worker>\d+) row=(?P<row>\d+) elapsed_s=(?P<elapsed>[0-9.]+)"
)


def _openpyxl_module() -> Any:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to execute IW51 workbook automation. Install dependencies with poetry."
        ) from exc
    return openpyxl


class _NullLogger:
    def info(self, message: str, *args: Any) -> None:
        return None

    def warning(self, message: str, *args: Any) -> None:
        return None

    def error(self, message: str, *args: Any) -> None:
        return None


def _pump_waiting_messages() -> None:
    try:
        import pythoncom  # type: ignore
    except Exception:
        return
    try:
        pythoncom.PumpWaitingMessages()
    except Exception:
        return


def wait_not_busy(session: Any, *, timeout_seconds: float = 30.0) -> None:
    _pump_waiting_messages()
    _sap_wait_not_busy(session, timeout_seconds=timeout_seconds)
    _pump_waiting_messages()


def _co_initialize(*, apartment_threaded: bool = False) -> tuple[Any | None, bool]:
    try:
        import pythoncom  # type: ignore
    except Exception:
        return None, False
    if apartment_threaded:
        pythoncom.CoInitializeEx(pythoncom.COINIT_APARTMENTTHREADED)
    else:
        pythoncom.CoInitialize()
    return pythoncom, True


def _get_sap_application() -> Any:
    try:
        import win32com.client  # type: ignore
    except Exception as exc:  # pragma: no cover - Windows only runtime path
        raise RuntimeError("pywin32 is required to resolve SAP GUI scripting application.") from exc
    return win32com.client.GetObject("SAPGUI").GetScriptingEngine


def _register_iw51_message_filter(*, retry_window_seconds: float, logger: Any) -> tuple[Any | None, Any | None]:
    try:
        import pythoncom  # type: ignore
        from win32com.server.util import wrap  # type: ignore
    except Exception:
        return None, None

    iid_message_filter = getattr(pythoncom, "IID_IMessageFilter", None)
    if iid_message_filter is None:
        return None, None

    class _Iw51MessageFilter:
        _public_methods_ = ["HandleInComingCall", "RetryRejectedCall", "MessagePending"]
        _com_interfaces_ = [iid_message_filter]

        def __init__(self, retry_window: float) -> None:
            self.retry_window = max(0.0, retry_window)

        def HandleInComingCall(self, dw_call_type: int, htask_caller: int, dw_tick_count: int, lp_interface_info: int) -> int:  # noqa: N802, ANN001
            return getattr(pythoncom, "SERVERCALL_ISHANDLED", 0)

        def RetryRejectedCall(self, htask_callee: int, dw_tick_count: int, dw_reject_type: int) -> int:  # noqa: N802, ANN001
            servercall_retrylater = getattr(pythoncom, "SERVERCALL_RETRYLATER", 2)
            if dw_reject_type != servercall_retrylater:
                return -1
            if dw_tick_count >= int(self.retry_window * 1000):
                return -1
            return 250

        def MessagePending(self, htask_callee: int, dw_tick_count: int, dw_pending_type: int) -> int:  # noqa: N802, ANN001
            _pump_waiting_messages()
            return getattr(pythoncom, "PENDINGMSG_WAITDEFPROCESS", 2)

    try:
        filter_impl = _Iw51MessageFilter(retry_window_seconds)
        wrapped = wrap(filter_impl)
        previous = pythoncom.CoRegisterMessageFilter(wrapped)
        logger.info(
            "IW51 COM message filter registered retry_window_s=%.1f",
            retry_window_seconds,
        )
        return previous, wrapped
    except Exception as exc:  # pragma: no cover - Windows only runtime path
        logger.warning("IW51 could not register COM message filter error=%s", exc)
        return None, None


def _unregister_iw51_message_filter(previous_filter: Any | None, logger: Any) -> None:
    try:
        import pythoncom  # type: ignore
    except Exception:
        return
    try:
        pythoncom.CoRegisterMessageFilter(previous_filter)
    except Exception as exc:  # pragma: no cover - Windows only runtime path
        logger.warning("IW51 could not restore previous COM message filter error=%s", exc)


def _marshal_sap_application(count: int, logger: Any) -> list[Any]:
    import pythoncom  # type: ignore
    import win32com.client  # type: ignore

    logger.info("IW51 marshal: resolving SAP GUI scripting engine via GetObject('SAPGUI') count=%s", count)
    sap_gui = win32com.client.GetObject("SAPGUI")
    application = sap_gui.GetScriptingEngine
    logger.info("IW51 marshal: scripting engine acquired, creating %s interthread streams", count)
    streams: list[Any] = []
    for index in range(count):
        stream = pythoncom.CoMarshalInterThreadInterfaceInStream(
            pythoncom.IID_IDispatch,
            application._oleobj_,
        )
        streams.append(stream)
        logger.info("IW51 marshal: stream %s/%s created", index + 1, count)
    return streams


def _unmarshal_sap_application(stream: Any, logger: Any, worker_index: int) -> Any:
    import pythoncom  # type: ignore
    import win32com.client  # type: ignore

    logger.info("IW51 unmarshal: worker=%s releasing stream into thread-local IDispatch", worker_index)
    dispatch = pythoncom.CoGetInterfaceAndReleaseStream(stream, pythoncom.IID_IDispatch)
    application = win32com.client.Dispatch(dispatch)
    logger.info("IW51 unmarshal: worker=%s SAP application proxy ready", worker_index)
    return application


def _normalize_header(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or "").strip())
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    return "".join(ch.lower() for ch in ascii_text if ch.isalnum())


def _normalize_transaction_code(value: str) -> str:
    code = str(value or "").strip()
    if not code:
        raise RuntimeError("IW51 transaction code is empty.")
    if code.startswith("/"):
        return code
    return f"/n{code}"


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "f").rstrip("0").rstrip(".")
    return str(value).strip()


def _has_iw51_header_row(sheet: Any) -> tuple[dict[str, int], int]:
    header_index_by_key: dict[str, int] = {}
    feito_column_index = 0
    for column_index, cell in enumerate(sheet[1], start=1):
        header_key = _normalize_header(cell.value)
        if not header_key:
            continue
        header_index_by_key[header_key] = column_index
        if header_key == "feito":
            feito_column_index = column_index
    return header_index_by_key, feito_column_index


def _looks_like_headerless_iw51_sheet(sheet: Any) -> bool:
    pn = _cell_to_text(sheet.cell(row=1, column=1).value)
    instalacao = _cell_to_text(sheet.cell(row=1, column=2).value)
    tipologia = _cell_to_text(sheet.cell(row=1, column=3).value)
    return bool(pn and instalacao and tipologia)


def _read_text(item: Any) -> str:
    for attr_name in ("Text", "text"):
        try:
            return str(getattr(item, attr_name, "") or "")
        except Exception:
            continue
    return ""


def _read_session_id(session: Any) -> str:
    for attr_name in ("Id", "id"):
        try:
            value = str(getattr(session, attr_name, "") or "").strip()
        except Exception:
            continue
        if value:
            return value
    return ""


def _read_active_window(session: Any) -> Any | None:
    for attr_name in ("ActiveWindow", "activeWindow"):
        try:
            window = getattr(session, attr_name)
        except Exception:
            continue
        if window is not None:
            return window
    return None


def _read_window_type(window: Any) -> str:
    if window is None:
        return ""
    for attr_name in ("Type", "type"):
        try:
            value = str(getattr(window, attr_name, "") or "").strip()
        except Exception:
            continue
        if value:
            return value
    return ""


def _visible_controls(*, session: Any, limit: int = 40) -> list[str]:
    try:
        import importlib

        compat = importlib.import_module("sap_gui_export_compat")
        return list(compat._collect_visible_control_ids(session=session, limit=limit))
    except Exception:
        return []


def _connection_from_session(session: Any) -> Any:
    for attr_name in ("Parent", "parent"):
        try:
            parent = getattr(session, attr_name)
        except Exception:
            continue
        if parent is not None:
            return parent
    raise RuntimeError("Could not resolve SAP connection from base session.")


def _list_connection_sessions(connection: Any, *, limit: int = 12) -> list[Any]:
    sessions_coll = getattr(connection, "Sessions", None)
    if sessions_coll is not None:
        try:
            count = int(getattr(sessions_coll, "Count", 0))
            return [sessions_coll(i) for i in range(min(count, limit))]
        except Exception:
            pass
    sessions: list[Any] = []
    for index in range(limit):
        try:
            sessions.append(connection.Children(index))
        except Exception:
            break
    return sessions


def _list_application_connections(application: Any, *, limit: int = 8) -> list[Any]:
    connections: list[Any] = []
    for index in range(limit):
        try:
            connections.append(application.Children(index))
        except Exception:
            break
    return connections


def _session_identity(session: Any) -> str:
    session_id = _read_session_id(session)
    if session_id:
        return session_id
    return f"<session-object:{id(session)}>"


def _is_session_disconnected_error(exc: Exception) -> bool:
    text = str(exc).casefold()
    return (
        "desconectado de seus clientes" in text
        or "called object was disconnected" in text
        or "falha na chamada de procedimento remoto" in text
        or "remote procedure call failed" in text
        or "rpc server is unavailable" in text
        or "o servidor rpc nao esta disponivel" in text
        or "o servidor rpc não está disponível" in text
    )


def _is_systemic_iw51_error(exc: Exception) -> bool:
    text = str(exc).casefold()
    return (
        _is_session_disconnected_error(exc)
        or "could not resolve sap gui element" in text
        or "the control could not be found by id" in text
        or "sintaxe inválida" in text
        or "sintaxe invalida" in text
        or "enumerator of the collection cannot find an element with the specified index" in text
        or "cannot find an element with the specified index" in text
    )


def _is_terminal_iw51_business_error(exc: Exception) -> bool:
    text = " ".join(str(exc).casefold().split())
    terminal_markers = (
        "instalacao invalida",
        "instalação inválida",
        "instalacao nao encontrada",
        "instalação não encontrada",
        "pn invalido",
        "pn inválido",
        "pn inexistente",
        "tipologia invalida",
        "tipologia inválida",
        "campo obrigatorio",
        "campo obrigatório",
        "valor invalido",
        "valor inválido",
    )
    return any(marker in text for marker in terminal_markers)


@dataclass(frozen=True)
class Iw51Settings:
    demandante: str
    workbook_path: Path
    sheet_name: str
    inter_item_sleep_seconds: float
    max_rows_per_run: int
    notification_type: str = _IW51_FIXED_QMART
    reference_notification_number: str = _IW51_FIXED_QWRNUM
    session_count: int = 3
    execution_mode: str = "interleaved"
    post_login_wait_seconds: float = 6.0
    workbook_sync_batch_size: int = 25
    circuit_breaker_fast_fail_threshold: int = 10
    circuit_breaker_slow_fail_threshold: int = 30
    per_step_timeout_seconds: float = 30.0
    session_recovery_mode: str = "soft"
    worker_stagger_seconds: float = 1.5
    worker_timeout_seconds: float = 3600.0
    worker_start_stagger_seconds: float = 2.0
    worker_heartbeat_seconds: float = 5.0
    worker_item_timeout_seconds: float = 180.0
    worker_restart_limit: int = 5
    worker_restart_backoff_seconds: float = 15.0
    session_rebuild_backoff_seconds: float = 30.0
    message_filter_retry_window_seconds: float = 30.0


@dataclass(frozen=True)
class Iw51WorkItem:
    row_index: int
    pn: str
    instalacao: str
    tipologia: str


@dataclass(frozen=True)
class Iw51ItemResult:
    row_index: int
    pn: str
    instalacao: str
    tipologia: str
    worker_index: int
    elapsed_seconds: float


@dataclass(frozen=True)
class Iw51SessionLocator:
    connection_index: int
    session_index: int


@dataclass
class Iw51WorkerState:
    worker_index: int
    session_id: str
    status: str = "idle"
    items_total: int = 0
    items_processed: int = 0
    items_ok: int = 0
    items_failed: int = 0
    last_ok_row_index: int = 0
    current_row_index: int = 0
    consecutive_failures: int = 0
    elapsed_seconds: float = 0.0
    first_item_started_at: float = 0.0
    last_item_finished_at: float = 0.0
    avg_item_seconds: float = 0.0

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass(frozen=True)
class Iw51LedgerState:
    success_rows: set[int]
    terminal_rows: set[int]
    rejected_rows: set[int]
    failed_terminal_rows: set[int]
    retriable_failed_rows: set[int]
    row_status_by_index: dict[int, str]
    last_entry_by_index: dict[int, dict[str, str]]
    total_entries: int

    @property
    def unique_rows(self) -> int:
        return len(self.row_status_by_index)


@dataclass(frozen=True)
class Iw51Manifest:
    run_id: str
    demandante: str
    workbook_path: str
    source_workbook_path: str
    working_workbook_path: str
    progress_ledger_path: str
    sheet_name: str
    processed_rows: int
    successful_rows: int
    skipped_rows: int
    rejected_rows: int
    status: str
    log_path: str
    manifest_path: str
    supervisor_mode: str = "none"
    recovered_from_log: bool = False
    worker_restarts: dict[str, int] = field(default_factory=dict)
    session_rebuilds: dict[str, int] = field(default_factory=dict)
    heartbeat_lag_seconds: dict[str, float] = field(default_factory=dict)
    remaining_rows: int = 0
    failed_rows: list[dict[str, Any]] = field(default_factory=list)
    worker_states: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class _Iw51ProcessWorkerRuntime:
    worker_index: int
    locator: Iw51SessionLocator
    group: list[Iw51WorkItem]
    process: Any | None = None
    next_item_index: int = 0
    current_row_index: int = 0
    current_item_started_at: float = 0.0
    last_heartbeat_at: float = 0.0
    restart_count: int = 0
    session_rebuild_count: int = 0
    done: bool = False


def load_iw51_settings(
    *,
    config: dict[str, Any],
    config_path: Path,
    demandante: str,
) -> Iw51Settings:
    iw51_cfg = config.get("iw51", {})
    demandante_name = str(demandante or "").strip().upper() or "DANI"
    if not isinstance(iw51_cfg, dict):
        raise RuntimeError("Missing iw51 section in config JSON.")
    demandantes_cfg = iw51_cfg.get("demandantes", {})
    if not isinstance(demandantes_cfg, dict):
        raise RuntimeError("Invalid iw51.demandantes section in config JSON.")
    profile = demandantes_cfg.get(demandante_name)
    if not isinstance(profile, dict):
        raise RuntimeError(f"Missing IW51 config for demandante={demandante_name}.")

    workbook_path = Path(str(profile.get("workbook_path", "")).strip())
    if not workbook_path.is_absolute():
        workbook_path = (config_path.expanduser().resolve().parent / workbook_path).resolve()
    execution_mode = str(profile.get("execution_mode", "interleaved")).strip().lower() or "interleaved"
    if execution_mode not in _IW51_EXECUTION_MODES:
        allowed = ", ".join(sorted(_IW51_EXECUTION_MODES))
        raise RuntimeError(
            f"Invalid iw51 execution_mode='{execution_mode}' for demandante={demandante_name}. "
            f"Allowed values: {allowed}."
        )
    return Iw51Settings(
        demandante=demandante_name,
        workbook_path=workbook_path,
        sheet_name=str(profile.get("sheet_name", "Macro1")).strip() or "Macro1",
        inter_item_sleep_seconds=float(profile.get("inter_item_sleep_seconds", 0.0)),
        max_rows_per_run=max(1, int(profile.get("max_rows_per_run", 3000) or 3000)),
        notification_type=str(profile.get("notification_type", _IW51_FIXED_QMART)).strip() or _IW51_FIXED_QMART,
        reference_notification_number=(
            str(profile.get("reference_notification_number", _IW51_FIXED_QWRNUM)).strip()
            or _IW51_FIXED_QWRNUM
        ),
        session_count=max(1, int(profile.get("session_count", 3) or 3)),
        execution_mode=execution_mode,
        post_login_wait_seconds=float(profile.get("post_login_wait_seconds", 6.0)),
        workbook_sync_batch_size=max(1, int(profile.get("workbook_sync_batch_size", 25) or 25)),
        circuit_breaker_fast_fail_threshold=max(
            1,
            int(profile.get("circuit_breaker_fast_fail_threshold", 10) or 10),
        ),
        circuit_breaker_slow_fail_threshold=max(
            1,
            int(profile.get("circuit_breaker_slow_fail_threshold", 30) or 30),
        ),
        per_step_timeout_seconds=float(profile.get("per_step_timeout_seconds", 30.0)),
        session_recovery_mode=str(profile.get("session_recovery_mode", "soft")).strip().lower() or "soft",
        worker_stagger_seconds=max(0.0, float(profile.get("worker_stagger_seconds", 1.5) or 1.5)),
        worker_timeout_seconds=max(1.0, float(profile.get("worker_timeout_seconds", 3600.0) or 3600.0)),
        worker_start_stagger_seconds=max(
            0.0,
            float(profile.get("worker_start_stagger_seconds", 2.0) or 2.0),
        ),
        worker_heartbeat_seconds=max(1.0, float(profile.get("worker_heartbeat_seconds", 5.0) or 5.0)),
        worker_item_timeout_seconds=max(
            1.0,
            float(profile.get("worker_item_timeout_seconds", 180.0) or 180.0),
        ),
        worker_restart_limit=max(0, int(profile.get("worker_restart_limit", 5) or 5)),
        worker_restart_backoff_seconds=max(
            0.0,
            float(profile.get("worker_restart_backoff_seconds", 15.0) or 15.0),
        ),
        session_rebuild_backoff_seconds=max(
            0.0,
            float(profile.get("session_rebuild_backoff_seconds", 30.0) or 30.0),
        ),
        message_filter_retry_window_seconds=max(
            0.0,
            float(profile.get("message_filter_retry_window_seconds", 30.0) or 30.0),
        ),
    )


def _copy_working_workbook(
    *,
    source_path: Path,
    working_path: Path,
    logger: Any,
) -> bool:
    working_path.parent.mkdir(parents=True, exist_ok=True)
    if working_path.exists():
        logger.info("IW51 reusing working workbook path=%s", working_path)
        return False
    shutil.copy2(source_path, working_path)
    logger.info("IW51 working workbook created source=%s target=%s", source_path, working_path)
    return True


def _save_workbook_atomic(*, workbook: Any, output_path: Path) -> None:
    temp_path = output_path.with_name(f"{output_path.name}.tmp")
    workbook.save(temp_path)
    os.replace(temp_path, output_path)


def _load_iw51_ledger_state(ledger_path: Path) -> Iw51LedgerState:
    success_rows: set[int] = set()
    terminal_rows: set[int] = set()
    rejected_rows: set[int] = set()
    failed_terminal_rows: set[int] = set()
    retriable_failed_rows: set[int] = set()
    row_status_by_index: dict[int, str] = {}
    last_entry_by_index: dict[int, dict[str, str]] = {}
    total_entries = 0
    if not ledger_path.exists():
        return Iw51LedgerState(
            success_rows=success_rows,
            terminal_rows=terminal_rows,
            rejected_rows=rejected_rows,
            failed_terminal_rows=failed_terminal_rows,
            retriable_failed_rows=retriable_failed_rows,
            row_status_by_index=row_status_by_index,
            last_entry_by_index=last_entry_by_index,
            total_entries=total_entries,
        )

    with ledger_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            try:
                row_index = int(str(row.get("row_index", "")).strip())
            except Exception:
                continue
            status = str(row.get("status", "")).strip().lower()
            if not status:
                continue
            row_status_by_index[row_index] = status
            last_entry_by_index[row_index] = {
                field_name: str(row.get(field_name, ""))
                for field_name in _IW51_LEDGER_FIELDNAMES
            }
            total_entries += 1

    for row_index, status in row_status_by_index.items():
        if status == "success":
            success_rows.add(row_index)
            terminal_rows.add(row_index)
        elif status == "rejected":
            rejected_rows.add(row_index)
            terminal_rows.add(row_index)
        elif status == "failed_terminal":
            failed_terminal_rows.add(row_index)
            terminal_rows.add(row_index)
        elif status == "failed":
            retriable_failed_rows.add(row_index)

    return Iw51LedgerState(
        success_rows=success_rows,
        terminal_rows=terminal_rows,
        rejected_rows=rejected_rows,
        failed_terminal_rows=failed_terminal_rows,
        retriable_failed_rows=retriable_failed_rows,
        row_status_by_index=row_status_by_index,
        last_entry_by_index=last_entry_by_index,
        total_entries=total_entries,
    )


def _compact_iw51_ledger(ledger_path: Path) -> int:
    ledger_state = _load_iw51_ledger_state(ledger_path)
    if ledger_state.total_entries <= ledger_state.unique_rows:
        return 0

    temp_path = ledger_path.with_name(f"{ledger_path.name}.tmp")
    with temp_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=_IW51_LEDGER_FIELDNAMES, lineterminator="\n")
        writer.writeheader()
        for row_index in sorted(ledger_state.row_status_by_index):
            writer.writerow(ledger_state.last_entry_by_index[row_index])
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp_path, ledger_path)
    return ledger_state.total_entries - ledger_state.unique_rows


def append_iw51_progress_ledger(
    *,
    ledger_path: Path,
    rows: list[dict[str, Any]],
) -> None:
    if not rows:
        return
    ledger_path.parent.mkdir(parents=True, exist_ok=True)
    file_exists = ledger_path.exists()
    with ledger_path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=_IW51_LEDGER_FIELDNAMES,
            lineterminator="\n",
        )
        if not file_exists:
            writer.writeheader()
        for row in rows:
            writer.writerow(row)
        handle.flush()
        os.fsync(handle.fileno())


def _resolve_iw51_sheet_layout(sheet: Any) -> tuple[dict[str, int], int, int]:
    header_index_by_key, feito_column_index = _has_iw51_header_row(sheet)
    data_start_row = 2
    missing_headers = [header for header in _IW51_REQUIRED_HEADERS if header not in header_index_by_key]
    if missing_headers:
        if _looks_like_headerless_iw51_sheet(sheet):
            header_index_by_key = dict(_IW51_HEADERLESS_DEFAULT_COLUMNS)
            feito_column_index = _IW51_HEADERLESS_DONE_COLUMN
            data_start_row = 1
        else:
            raise RuntimeError(
                "IW51 workbook is missing required columns: " + ", ".join(sorted(missing_headers))
            )
    return header_index_by_key, feito_column_index, data_start_row


def _collect_iw51_workbook_done_rows(
    *,
    sheet: Any,
    header_index_by_key: dict[str, int],
    feito_column_index: int,
    data_start_row: int,
    ledger_terminal_rows: set[int],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_index in range(data_start_row, sheet.max_row + 1):
        pn = _cell_to_text(sheet.cell(row=row_index, column=header_index_by_key["pn"]).value)
        instalacao = _cell_to_text(sheet.cell(row=row_index, column=header_index_by_key["instalacao"]).value)
        tipologia = _cell_to_text(sheet.cell(row=row_index, column=header_index_by_key["tipologia"]).value)
        feito = _cell_to_text(sheet.cell(row=row_index, column=feito_column_index).value).upper()
        if not pn and not instalacao and not tipologia:
            continue
        if feito != _IW51_DONE_VALUE or row_index in ledger_terminal_rows:
            continue
        rows.append(
            {
                "row_index": row_index,
                "worker_index": 0,
                "pn": pn,
                "instalacao": instalacao,
                "tipologia": tipologia,
                "status": "success",
                "attempt": 0,
                "elapsed_seconds": 0.0,
                "error": "Imported from workbook FEITO state",
            }
        )
    return rows


def _count_iw51_workbook_done_rows(*, sheet: Any, feito_column_index: int, data_start_row: int) -> int:
    count = 0
    for row_index in range(data_start_row, sheet.max_row + 1):
        feito = _cell_to_text(sheet.cell(row=row_index, column=feito_column_index).value).upper()
        if feito == _IW51_DONE_VALUE:
            count += 1
    return count


def _write_iw51_ledger_snapshot(*, ledger_path: Path, rows: list[dict[str, Any]]) -> None:
    ledger_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = ledger_path.with_name(f"{ledger_path.name}.tmp")
    with temp_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=_IW51_LEDGER_FIELDNAMES, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
        handle.flush()
        os.fsync(handle.fileno())
    os.replace(temp_path, ledger_path)


def _recover_iw51_success_rows_from_log(log_path: Path) -> list[dict[str, Any]]:
    completed_by_row: dict[int, dict[str, Any]] = {}
    elapsed_by_row: dict[int, float] = {}
    for line in log_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        completed_match = _IW51_LOG_COMPLETED_RE.search(line)
        if completed_match:
            row_index = int(completed_match.group("row"))
            completed_by_row[row_index] = {
                "row_index": row_index,
                "worker_index": int(completed_match.group("worker")),
                "pn": completed_match.group("pn"),
                "instalacao": completed_match.group("instalacao"),
                "tipologia": completed_match.group("tipologia"),
            }
            continue

        ok_match = _IW51_LOG_OK_RE.search(line)
        if ok_match:
            elapsed_by_row[int(ok_match.group("row"))] = float(ok_match.group("elapsed"))

    recovered_rows: list[dict[str, Any]] = []
    for row_index in sorted(completed_by_row):
        row = completed_by_row[row_index]
        recovered_rows.append(
            {
                **row,
                "status": "success",
                "attempt": 1,
                "elapsed_seconds": round(elapsed_by_row.get(row_index, 0.0), 3),
                "error": "Recovered from sap_session.log",
            }
        )
    return recovered_rows


def recover_iw51_run_artifacts_from_log(
    *,
    run_root: Path,
    workbook_path: Path | None = None,
    sheet_name: str = "Macro1",
) -> dict[str, Any]:
    log_path = run_root / "logs" / "sap_session.log"
    ledger_path = run_root / "iw51" / "iw51_progress.csv"
    manifest_path = run_root / "iw51" / "iw51_manifest.json"
    if not log_path.exists():
        raise RuntimeError(f"IW51 recovery log not found: {log_path}")

    recovered_rows = _recover_iw51_success_rows_from_log(log_path)
    _write_iw51_ledger_snapshot(ledger_path=ledger_path, rows=recovered_rows)

    workbook_rows_synced = 0
    if workbook_path is not None and workbook_path.exists():
        openpyxl = _openpyxl_module()
        workbook = openpyxl.load_workbook(workbook_path, keep_vba=True)
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f"Sheet '{sheet_name}' not found in workbook {workbook_path}.")
        sheet = workbook[sheet_name]
        _, feito_column_index, _ = _resolve_iw51_sheet_layout(sheet)
        _sync_workbook_done_rows(
            workbook=workbook,
            sheet=sheet,
            feito_column_index=feito_column_index,
            workbook_path=workbook_path,
            row_indices={int(row["row_index"]) for row in recovered_rows},
            logger=_NullLogger(),
        )
        workbook_rows_synced = len(recovered_rows)

    manifest_updates: dict[str, Any] = {}
    if manifest_path.exists():
        manifest_data = json.loads(manifest_path.read_text(encoding="utf-8"))
        manifest_data["successful_rows"] = len(recovered_rows)
        manifest_data["status"] = "partial" if manifest_data.get("failed_rows") else "success"
        manifest_data["recovered_from_log"] = True
        manifest_path.write_text(json.dumps(manifest_data, ensure_ascii=False, indent=2), encoding="utf-8")
        manifest_updates = {
            "successful_rows": manifest_data["successful_rows"],
            "status": manifest_data["status"],
        }

    return {
        "log_path": str(log_path),
        "ledger_path": str(ledger_path),
        "manifest_path": str(manifest_path),
        "recovered_success_rows": len(recovered_rows),
        "workbook_rows_synced": workbook_rows_synced,
        **manifest_updates,
    }


def load_iw51_work_items(
    *,
    workbook_path: Path,
    sheet_name: str,
    max_rows: int | None = None,
    completed_row_indices: set[int] | None = None,
) -> tuple[Any, Any, int, list[Iw51WorkItem], int, list[dict[str, Any]]]:
    openpyxl = _openpyxl_module()
    workbook = openpyxl.load_workbook(workbook_path, keep_vba=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"Sheet '{sheet_name}' not found in workbook {workbook_path}.")
    sheet = workbook[sheet_name]

    header_index_by_key, feito_column_index, data_start_row = _resolve_iw51_sheet_layout(sheet)

    if feito_column_index == 0 and data_start_row == 2:
        feito_column_index = sheet.max_column + 1
        sheet.cell(row=1, column=feito_column_index, value=_IW51_DONE_HEADER)

    completed_rows = completed_row_indices or set()
    items: list[Iw51WorkItem] = []
    skipped_rows = 0
    rejected_rows: list[dict[str, Any]] = []
    for row_index in range(data_start_row, sheet.max_row + 1):
        pn = _cell_to_text(sheet.cell(row=row_index, column=header_index_by_key["pn"]).value)
        instalacao = _cell_to_text(sheet.cell(row=row_index, column=header_index_by_key["instalacao"]).value)
        tipologia = _cell_to_text(sheet.cell(row=row_index, column=header_index_by_key["tipologia"]).value)
        feito = _cell_to_text(sheet.cell(row=row_index, column=feito_column_index).value).upper()

        if not pn and not instalacao and not tipologia:
            continue
        if feito == _IW51_DONE_VALUE or row_index in completed_rows:
            skipped_rows += 1
            continue
        if not pn or not instalacao or not tipologia:
            rejected_rows.append(
                {
                    "worker_index": 0,
                    "row_index": row_index,
                    "pn": pn,
                    "instalacao": instalacao,
                    "tipologia": tipologia,
                    "elapsed_seconds": 0.0,
                    "attempt": 0,
                    "fast_fail": False,
                    "systemic": False,
                    "status": "rejected",
                    "error": (
                        f"IW51 workbook row {row_index} is incomplete. "
                        "Expected pn, instalacao and tipologia."
                    ),
                }
            )
            continue

        items.append(
            Iw51WorkItem(
                row_index=row_index,
                pn=pn,
                instalacao=instalacao,
                tipologia=tipologia,
            )
        )
        if max_rows is not None and max_rows > 0 and len(items) >= max_rows:
            break

    return workbook, sheet, feito_column_index, items, skipped_rows, rejected_rows


def _read_status_bar_text(session: Any) -> str:
    for item_id in ("wnd[0]/sbar/pane[0]", "wnd[0]/sbar"):
        try:
            return _read_text(session.findById(item_id)).strip()
        except Exception:
            continue
    return ""


def _log_iw51_snapshot(*, session: Any, logger: Any, phase: str, worker_index: int, item: Iw51WorkItem | None) -> None:
    active_window = _read_active_window(session)
    logger.info(
        "IW51 snapshot phase=%s worker=%s row=%s session_id=%s active_window_type=%s status_bar=%s visible_controls=%s",
        phase,
        worker_index,
        item.row_index if item is not None else "<empty>",
        _read_session_id(session) or "<empty>",
        _read_window_type(active_window) or "<empty>",
        _read_status_bar_text(session) or "<empty>",
        _visible_controls(session=session, limit=20),
    )


def _wait_for_control(
    *,
    session: Any,
    ids: list[str],
    timeout_seconds: float,
    logger: Any,
    description: str,
    worker_index: int,
    item: Iw51WorkItem,
) -> Any:
    deadline = time.monotonic() + max(1.0, timeout_seconds)
    last_error: Exception | None = None
    while time.monotonic() < deadline:
        try:
            _, control = resolve_first_existing_with_id(session, ids)
            return control
        except Exception as exc:
            last_error = exc
            if _is_session_disconnected_error(exc):
                break
            _dismiss_popup_if_present(session, logger)
            time.sleep(0.2)
    _log_iw51_snapshot(
        session=session,
        logger=logger,
        phase=f"wait_control_timeout.{description}",
        worker_index=worker_index,
        item=item,
    )
    raise RuntimeError(
        f"IW51 could not resolve control for {description} worker={worker_index} row={item.row_index}: {last_error}"
    )


def _wait_session_ready(session: Any, *, timeout_seconds: float, logger: Any) -> None:
    deadline = time.monotonic() + min(timeout_seconds, 30.0)
    while time.monotonic() < deadline:
        try:
            if getattr(session, "Busy", False):
                time.sleep(0.3)
                continue
            main_window = session.findById(_IW51_MAIN_WINDOW_ID)
            okcd = session.findById(_IW51_OKCODE_ID)
            active_window = _read_active_window(session)
            active_window_type = _read_window_type(active_window)
            if active_window is not None and active_window_type != "GuiModalWindow" and main_window and okcd:
                return
        except Exception:
            pass
        time.sleep(0.3)
    logger.warning(
        "IW51 session readiness timed out session_id=%s active_window_type=%s — proceeding anyway",
        _read_session_id(session),
        _read_window_type(_read_active_window(session)) or "<empty>",
    )


def _dismiss_popup_if_present(session: Any, logger: Any) -> bool:
    try:
        popup = session.findById("wnd[1]")
    except Exception:
        return False

    popup_type = _read_window_type(popup) or "<unknown>"
    popup_text = _read_text(popup)
    logger.warning(
        "IW51 popup detected type=%s text=%s session_id=%s",
        popup_type,
        popup_text or "<empty>",
        _read_session_id(session) or "<empty>",
    )

    try:
        popup.sendVKey(0)
        logger.info("IW51 popup dismissed via Enter type=%s", popup_type)
        return True
    except Exception:
        pass

    for item_id in ("wnd[1]/tbar[0]/btn[0]", "wnd[1]/usr/btnSPOP-OPTION1", "wnd[1]/usr/btnBUTTON_1"):
        try:
            session.findById(item_id).press()
            logger.info("IW51 popup dismissed via button id=%s", item_id)
            return True
        except Exception:
            continue
    logger.warning("IW51 popup could not be dismissed automatically type=%s", popup_type)
    return False


def ensure_sap_sessions(
    *,
    base_session: Any,
    session_count: int,
    logger: Any,
    wait_timeout_seconds: float,
) -> list[Any]:
    if session_count <= 1:
        return [base_session]
    connection = _connection_from_session(base_session)
    sessions = [base_session]
    known_identities = {_session_identity(base_session)}
    main_window = base_session.findById(_IW51_MAIN_WINDOW_ID)
    try:
        main_window.maximize()
    except Exception:
        pass
    logger.info("IW51 registered SAP session slot=1 session_id=%s", _read_session_id(base_session) or "<empty>")
    while len(sessions) < session_count:
        next_slot = len(sessions) + 1
        logger.info("IW51 opening additional SAP session slot=%s", next_slot)
        try:
            main_window.sendVKey(0)
            wait_not_busy(base_session, timeout_seconds=wait_timeout_seconds)
        except Exception:
            pass
        create_session = getattr(base_session, "CreateSession", None) or getattr(base_session, "createSession", None)
        if callable(create_session):
            create_session()
        else:
            set_text(base_session, _IW51_OKCODE_ID, "/o")
            base_session.findById(_IW51_MAIN_WINDOW_ID).sendVKey(0)
        deadline = time.monotonic() + max(5.0, wait_timeout_seconds)
        new_session: Any | None = None
        while time.monotonic() < deadline:
            current_sessions = _list_connection_sessions(connection)
            for candidate in current_sessions:
                identity = _session_identity(candidate)
                if identity in known_identities:
                    continue
                new_session = candidate
                known_identities.add(identity)
                break
            if new_session is not None:
                _wait_session_ready(new_session, timeout_seconds=wait_timeout_seconds, logger=logger)
                sessions.append(new_session)
                try:
                    new_session.findById(_IW51_MAIN_WINDOW_ID).maximize()
                except Exception:
                    pass
                logger.info(
                    "IW51 registered SAP session slot=%s session_id=%s",
                    len(sessions),
                    _read_session_id(new_session) or "<empty>",
                )
                break
            time.sleep(0.25)
        else:
            raise RuntimeError(f"Could not open SAP session slot={next_slot} for IW51 flow.")
    return sessions


def prepare_iw51_sessions(
    *,
    base_session: Any,
    settings: Iw51Settings,
    logger: Any,
) -> list[Any]:
    if settings.post_login_wait_seconds > 0:
        logger.info(
            "IW51 waiting after authenticated login seconds=%.1f before opening SAP sessions",
            settings.post_login_wait_seconds,
        )
        time.sleep(settings.post_login_wait_seconds)
    return ensure_sap_sessions(
        base_session=base_session,
        session_count=settings.session_count,
        logger=logger,
        wait_timeout_seconds=settings.per_step_timeout_seconds,
    )


def _session_locator_from_session(session: Any) -> Iw51SessionLocator:
    session_id = _read_session_id(session)
    fragments = session_id.split("/")
    try:
        connection_part = next(fragment for fragment in fragments if fragment.startswith("con["))
        session_part = next(fragment for fragment in fragments if fragment.startswith("ses["))
        return Iw51SessionLocator(
            connection_index=int(connection_part[4:-1]),
            session_index=int(session_part[4:-1]),
        )
    except Exception as exc:
        raise RuntimeError(f"Could not determine IW51 session locator from session id: {session_id or '<empty>'}") from exc


def _reattach_iw51_session(
    *,
    locator: Iw51SessionLocator,
    logger: Any,
    application: Any | None = None,
) -> Any:
    if application is None:
        try:
            import win32com.client  # type: ignore
        except Exception as exc:  # pragma: no cover - Windows only runtime path
            raise RuntimeError("pywin32 is required to reattach SAP sessions for IW51 flow.") from exc
        application = win32com.client.GetObject("SAPGUI").GetScriptingEngine

    try:
        connection = application.Children(locator.connection_index)
    except Exception as exc:
        raise RuntimeError(
            f"IW51 could not reattach SAP connection index={locator.connection_index}: {exc}"
        ) from exc

    target_id = f"/app/con[{locator.connection_index}]/ses[{locator.session_index}]"
    sessions_coll = getattr(connection, "Sessions", None)
    if sessions_coll is not None:
        try:
            count = int(getattr(sessions_coll, "Count", 0))
            for index in range(count):
                candidate = sessions_coll(index)
                if _read_session_id(candidate) == target_id:
                    logger.info("IW51 reattached session by id=%s (Sessions index=%s)", target_id, index)
                    return candidate
        except Exception:
            pass

    try:
        session = connection.Children(locator.session_index)
        logger.info(
            "IW51 reattached SAP session (Children fallback) con=%s ses=%s session_id=%s",
            locator.connection_index,
            locator.session_index,
            _read_session_id(session) or "<empty>",
        )
        return session
    except Exception:
        pass

    existing_sessions = _list_connection_sessions(connection)
    if not existing_sessions:
        raise RuntimeError(
            f"IW51 could not reattach session con={locator.connection_index} ses={locator.session_index}: "
            "connection has no active sessions."
        )

    logger.warning(
        "IW51 session slot missing; attempting to recreate slot con=%s ses=%s target_id=%s active_sessions=%s",
        locator.connection_index,
        locator.session_index,
        target_id,
        [_read_session_id(session) or "<empty>" for session in existing_sessions],
    )

    anchor_session = existing_sessions[0]
    anchor_window = anchor_session.findById(_IW51_MAIN_WINDOW_ID)
    try:
        anchor_window.sendVKey(0)
        wait_not_busy(anchor_session, timeout_seconds=5.0)
    except Exception:
        pass

    deadline = time.monotonic() + 20.0
    while time.monotonic() < deadline:
        current_sessions = _list_connection_sessions(connection)
        if len(current_sessions) > locator.session_index:
            session = current_sessions[locator.session_index]
            logger.info(
                "IW51 recreated missing session slot con=%s ses=%s session_id=%s",
                locator.connection_index,
                locator.session_index,
                _read_session_id(session) or "<empty>",
            )
            return session

        create_session = getattr(anchor_session, "CreateSession", None) or getattr(anchor_session, "createSession", None)
        if callable(create_session):
            create_session()
        else:
            set_text(anchor_session, _IW51_OKCODE_ID, "/o")
            anchor_window.sendVKey(0)
        time.sleep(0.5)

    raise RuntimeError(
        f"IW51 could not recreate session slot con={locator.connection_index} ses={locator.session_index}."
    )


def _reset_session_state_soft(*, session: Any, settings: Iw51Settings, logger: Any) -> None:
    _dismiss_popup_if_present(session, logger)
    main_window = session.findById(_IW51_MAIN_WINDOW_ID)
    set_text(session, _IW51_OKCODE_ID, "/n")
    logger.info("IW51 soft session reset transaction=/n session_id=%s", _read_session_id(session) or "<empty>")
    main_window.sendVKey(0)
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    _dismiss_popup_if_present(session, logger)


def _prepare_iw51_item_surface(
    *,
    session: Any,
    settings: Iw51Settings,
    logger: Any,
    worker_index: int,
    item: Iw51WorkItem,
) -> None:
    _dismiss_popup_if_present(session, logger)
    _reset_session_state_soft(session=session, settings=settings, logger=logger)
    main_window = session.findById(_IW51_MAIN_WINDOW_ID)
    transaction_code = _normalize_transaction_code(_IW51_TRANSACTION_CODE)
    set_text(session, _IW51_OKCODE_ID, transaction_code)
    logger.info(
        "IW51 entering transaction worker=%s row=%s transaction=%s",
        worker_index,
        item.row_index,
        transaction_code,
    )
    main_window.sendVKey(0)
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    _dismiss_popup_if_present(session, logger)
    _wait_for_control(
        session=session,
        ids=_IW51_ENTRY_FIELD_IDS,
        timeout_seconds=settings.per_step_timeout_seconds,
        logger=logger,
        description="iw51.entry_screen",
        worker_index=worker_index,
        item=item,
    )


def execute_iw51_item(
    *,
    session: Any,
    item: Iw51WorkItem,
    settings: Iw51Settings,
    logger: Any,
    worker_index: int = 1,
) -> None:
    main_window = session.findById(_IW51_MAIN_WINDOW_ID)
    _prepare_iw51_item_surface(
        session=session,
        settings=settings,
        logger=logger,
        worker_index=worker_index,
        item=item,
    )

    logger.info("IW51 step fill_header worker=%s row=%s", worker_index, item.row_index)
    set_text(session, _IW51_QMART_ID, settings.notification_type)
    qwrnum_item = session.findById(_IW51_QWRNUM_ID)
    qwrnum_item.text = settings.reference_notification_number
    try:
        qwrnum_item.setFocus()
    except Exception:
        pass
    try:
        qwrnum_item.caretPosition = len(settings.reference_notification_number)
    except Exception:
        pass
    main_window.sendVKey(0)
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    _dismiss_popup_if_present(session, logger)

    logger.info("IW51 step fill_instalacao worker=%s row=%s", worker_index, item.row_index)
    session.findById(_IW51_TAB_19_ID).select()
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    set_text(session, _IW51_TPLNR_ID, "")
    instalacao_item = session.findById(_IW51_INSTALACAO_ID)
    instalacao_item.text = item.instalacao
    try:
        instalacao_item.setFocus()
    except Exception:
        pass
    try:
        instalacao_item.caretPosition = len(item.instalacao)
    except Exception:
        pass
    main_window.sendVKey(0)
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    _dismiss_popup_if_present(session, logger)

    logger.info("IW51 step fill_pn_tipologia worker=%s row=%s", worker_index, item.row_index)
    session.findById(_IW51_TAB_01_ID).select()
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    pn_item = session.findById(_IW51_PN_ID)
    pn_item.text = item.pn
    try:
        pn_item.setFocus()
    except Exception:
        pass
    try:
        pn_item.caretPosition = len(item.pn)
    except Exception:
        pass
    main_window.sendVKey(0)
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    _dismiss_popup_if_present(session, logger)

    tipologia_item = session.findById(_IW51_TIPOLOGIA_ID)
    tipologia_item.text = item.tipologia
    try:
        tipologia_item.setFocus()
    except Exception:
        pass
    try:
        tipologia_item.caretPosition = len(item.tipologia)
    except Exception:
        pass

    logger.info("IW51 step set_status worker=%s row=%s step=1", worker_index, item.row_index)
    session.findById(_IW51_STATUS_BUTTON_ID).press()
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    first_status = session.findById(_IW51_STATUS_FIRST_ID)
    set_selected(first_status, True)
    try:
        first_status.setFocus()
    except Exception:
        pass
    session.findById(_IW51_STATUS_CONFIRM_ID).press()
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    _dismiss_popup_if_present(session, logger)

    logger.info("IW51 step set_status worker=%s row=%s step=2", worker_index, item.row_index)
    session.findById(_IW51_STATUS_BUTTON_ID).press()
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    second_status = session.findById(_IW51_STATUS_SECOND_ID)
    set_selected(second_status, True)
    try:
        second_status.setFocus()
    except Exception:
        pass
    session.findById(_IW51_STATUS_CONFIRM_ID).press()
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    _dismiss_popup_if_present(session, logger)

    logger.info("IW51 step save worker=%s row=%s", worker_index, item.row_index)
    session.findById(_IW51_SAVE_BUTTON_ID).press()
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    session.findById(_IW51_SAVE_CONFIRM_ID).press()
    wait_not_busy(session, timeout_seconds=settings.per_step_timeout_seconds)
    _dismiss_popup_if_present(session, logger)

    logger.info(
        "IW51 item completed worker=%s row=%s pn=%s instalacao=%s tipologia=%s",
        worker_index,
        item.row_index,
        item.pn,
        item.instalacao,
        item.tipologia,
    )


def _process_iw51_item_with_retry(
    *,
    worker_index: int,
    session: Any,
    session_locator: Iw51SessionLocator,
    item: Iw51WorkItem,
    settings: Iw51Settings,
    logger: Any,
    application: Any | None = None,
    serialize_com_calls: bool = True,
) -> tuple[Iw51ItemResult | None, dict[str, Any] | None, Any]:
    current_session = session
    last_error: Exception | None = None
    last_nonsystemic_error_signature = ""
    for attempt in range(1, 3):
        started_at = time.perf_counter()
        try:
            with _IW51_SAP_COM_LOCK if serialize_com_calls else nullcontext():
                if attempt > 1:
                    logger.info(
                        "IW51 reattach for retry worker=%s row=%s attempt=%s con=%s ses=%s",
                        worker_index,
                        item.row_index,
                        attempt,
                        session_locator.connection_index,
                        session_locator.session_index,
                    )
                    current_session = _reattach_iw51_session(
                        locator=session_locator,
                        logger=logger,
                        application=application,
                    )
                    _wait_session_ready(
                        current_session,
                        timeout_seconds=settings.per_step_timeout_seconds,
                        logger=logger,
                    )
                    if settings.session_recovery_mode == "soft":
                        _reset_session_state_soft(session=current_session, settings=settings, logger=logger)
                execute_iw51_item(
                    session=current_session,
                    item=item,
                    settings=settings,
                    logger=logger,
                    worker_index=worker_index,
                )
        except Exception as exc:
            elapsed_seconds = time.perf_counter() - started_at
            last_error = exc
            is_fast_fail = elapsed_seconds < _IW51_FAST_FAIL_THRESHOLD_SECONDS
            systemic = _is_systemic_iw51_error(exc)
            error_signature = " ".join(str(exc).lower().split())
            if attempt < 2 and systemic:
                logger.warning(
                    "IW51 systemic failure; retrying worker=%s row=%s attempt=%s elapsed_s=%.2f error=%s",
                    worker_index,
                    item.row_index,
                    attempt,
                    elapsed_seconds,
                    exc,
                )
                continue
            status = "failed"
            if not systemic and (
                _is_terminal_iw51_business_error(exc)
                or (error_signature and error_signature == last_nonsystemic_error_signature)
            ):
                status = "failed_terminal"
            if not systemic and error_signature:
                last_nonsystemic_error_signature = error_signature
            logger.error(
                "IW51 item FAILED worker=%s row=%s elapsed_s=%.2f attempt=%s fast_fail=%s systemic=%s error=%s",
                worker_index,
                item.row_index,
                elapsed_seconds,
                attempt,
                is_fast_fail,
                systemic,
                exc,
            )
            return None, {
                "worker_index": worker_index,
                "row_index": item.row_index,
                "pn": item.pn,
                "instalacao": item.instalacao,
                "tipologia": item.tipologia,
                "elapsed_seconds": round(elapsed_seconds, 3),
                "attempt": attempt,
                "fast_fail": is_fast_fail,
                "systemic": systemic,
                "status": status,
                "error": str(exc),
            }, current_session

        elapsed_seconds = time.perf_counter() - started_at
        logger.info(
            "IW51 item OK worker=%s row=%s elapsed_s=%.2f",
            worker_index,
            item.row_index,
            elapsed_seconds,
        )
        return Iw51ItemResult(
            row_index=item.row_index,
            pn=item.pn,
            instalacao=item.instalacao,
            tipologia=item.tipologia,
            worker_index=worker_index,
            elapsed_seconds=round(elapsed_seconds, 3),
        ), None, current_session
    raise RuntimeError(f"Unexpected IW51 retry fallthrough for row={item.row_index}: {last_error}")


def split_iw51_work_items_evenly(items: list[Iw51WorkItem], session_count: int) -> list[list[Iw51WorkItem]]:
    if session_count <= 0:
        raise ValueError("session_count must be greater than zero.")
    groups = [[] for _ in range(session_count)]
    for index, item in enumerate(items):
        groups[index % session_count].append(item)
    return [group for group in groups if group]


def _sync_workbook_done_rows(
    *,
    workbook: Any,
    sheet: Any,
    feito_column_index: int,
    workbook_path: Path,
    row_indices: set[int],
    logger: Any,
) -> None:
    if not row_indices:
        return
    for row_index in sorted(row_indices):
        sheet.cell(row=row_index, column=feito_column_index, value=_IW51_DONE_VALUE)
    _save_workbook_atomic(workbook=workbook, output_path=workbook_path)
    logger.info(
        "IW51 workbook sync completed rows=%s workbook=%s",
        len(row_indices),
        workbook_path,
    )


def _build_ledger_row_from_result(result: Iw51ItemResult) -> dict[str, Any]:
    return {
        "row_index": result.row_index,
        "worker_index": result.worker_index,
        "pn": result.pn,
        "instalacao": result.instalacao,
        "tipologia": result.tipologia,
        "status": "success",
        "attempt": 1,
        "elapsed_seconds": result.elapsed_seconds,
        "error": "",
    }


def _build_ledger_row_from_failure(failure: dict[str, Any]) -> dict[str, Any]:
    return {
        "row_index": failure.get("row_index", 0),
        "worker_index": failure.get("worker_index", 0),
        "pn": failure.get("pn", ""),
        "instalacao": failure.get("instalacao", ""),
        "tipologia": failure.get("tipologia", ""),
        "status": failure.get("status", "failed"),
        "attempt": failure.get("attempt", 0),
        "elapsed_seconds": failure.get("elapsed_seconds", 0.0),
        "error": failure.get("error", ""),
    }


def _run_iw51_interleaved_workers(
    *,
    groups: list[list[Iw51WorkItem]],
    session_locators: list[Iw51SessionLocator],
    sessions: list[Any],
    settings: Iw51Settings,
    logger: Any,
    worker_states: list[Iw51WorkerState],
    apply_worker_results: Any,
) -> None:
    logger.info(
        "IW51 starting interleaved multi-session processing workers=%s total_items=%s",
        len(groups),
        sum(len(group) for group in groups),
    )
    worker_started_at = time.perf_counter()
    positions = {worker_index: 0 for worker_index in range(1, len(groups) + 1)}
    results_count = {worker_index: 0 for worker_index in range(1, len(groups) + 1)}
    failures_count = {worker_index: 0 for worker_index in range(1, len(groups) + 1)}
    consecutive_failures = {worker_index: 0 for worker_index in range(1, len(groups) + 1)}
    consecutive_fast_fails = {worker_index: 0 for worker_index in range(1, len(groups) + 1)}
    next_ready_at = {worker_index: 0.0 for worker_index in range(1, len(groups) + 1)}
    completed_workers: set[int] = set()
    current_sessions = {worker_index: sessions[worker_index - 1] for worker_index in range(1, len(groups) + 1)}

    for worker_index, group in enumerate(groups, start=1):
        worker_state = worker_states[worker_index - 1]
        worker_state.status = "running"
        worker_state.items_total = len(group)
        worker_state.session_id = _read_session_id(current_sessions[worker_index])
        logger.info(
            "IW51 worker START worker=%s assigned_items=%s con=%s ses=%s session_id=%s",
            worker_index,
            len(group),
            session_locators[worker_index - 1].connection_index,
            session_locators[worker_index - 1].session_index,
            worker_state.session_id,
        )

    while len(completed_workers) < len(groups):
        progressed = False
        for worker_index, group in enumerate(groups, start=1):
            if worker_index in completed_workers:
                continue
            if time.monotonic() < next_ready_at[worker_index]:
                continue

            index = positions[worker_index]
            if index >= len(group):
                worker_state = worker_states[worker_index - 1]
                worker_elapsed = time.perf_counter() - worker_started_at
                if worker_state.status not in {"circuit_breaker", "failed"}:
                    worker_state.status = "completed"
                worker_state.elapsed_seconds = round(worker_elapsed, 3)
                if worker_state.items_processed > 0:
                    worker_state.avg_item_seconds = round(worker_elapsed / worker_state.items_processed, 3)
                worker_state.current_row_index = 0
                logger.info(
                    "IW51 worker DONE worker=%s ok=%s fail=%s total=%s elapsed_s=%.1f",
                    worker_index,
                    results_count[worker_index],
                    failures_count[worker_index],
                    len(group),
                    worker_elapsed,
                )
                completed_workers.add(worker_index)
                continue

            item = group[index]
            if index == 0 or (index + 1) % _IW51_PROGRESS_LOG_INTERVAL == 0:
                elapsed_so_far = time.perf_counter() - worker_started_at
                rate = results_count[worker_index] / elapsed_so_far if elapsed_so_far > 0 else 0.0
                logger.info(
                    "IW51 worker PROGRESS worker=%s index=%s/%s ok=%s fail=%s elapsed_s=%.1f rate=%.2f items/s",
                    worker_index,
                    index + 1,
                    len(group),
                    results_count[worker_index],
                    failures_count[worker_index],
                    elapsed_so_far,
                    rate,
                )

            worker_state = worker_states[worker_index - 1]
            worker_state.current_row_index = item.row_index
            if worker_state.first_item_started_at == 0.0:
                worker_state.first_item_started_at = time.time()
            result, failure, current_session = _process_iw51_item_with_retry(
                worker_index=worker_index,
                session=current_sessions[worker_index],
                session_locator=session_locators[worker_index - 1],
                item=item,
                settings=settings,
                logger=logger,
            )
            current_sessions[worker_index] = current_session
            positions[worker_index] += 1
            progressed = True

            if result is not None:
                apply_worker_results(worker_index, [result], [])
                results_count[worker_index] += 1
                consecutive_failures[worker_index] = 0
                consecutive_fast_fails[worker_index] = 0
                worker_state.items_processed += 1
                worker_state.items_ok += 1
                worker_state.last_ok_row_index = item.row_index
                worker_state.consecutive_failures = 0
                worker_state.last_item_finished_at = time.time()
                if settings.inter_item_sleep_seconds > 0:
                    next_ready_at[worker_index] = time.monotonic() + settings.inter_item_sleep_seconds
                continue

            assert failure is not None
            apply_worker_results(worker_index, [], [failure])
            failures_count[worker_index] += 1
            consecutive_failures[worker_index] += 1
            worker_state.items_processed += 1
            worker_state.items_failed += 1
            worker_state.consecutive_failures = consecutive_failures[worker_index]
            worker_state.last_item_finished_at = time.time()
            if failure.get("fast_fail", False):
                consecutive_fast_fails[worker_index] += 1
            else:
                consecutive_fast_fails[worker_index] = 0

            remaining_items = group[positions[worker_index] :]
            if consecutive_fast_fails[worker_index] >= settings.circuit_breaker_fast_fail_threshold:
                logger.error(
                    "IW51 worker CIRCUIT-BREAKER worker=%s consecutive_fast_fails=%s — aborting worker, skipping %s remaining items.",
                    worker_index,
                    consecutive_fast_fails[worker_index],
                    len(remaining_items),
                )
                worker_state.status = "circuit_breaker"
                skipped_failures = [
                    {
                        "worker_index": worker_index,
                        "row_index": skip_item.row_index,
                        "pn": skip_item.pn,
                        "instalacao": skip_item.instalacao,
                        "tipologia": skip_item.tipologia,
                        "elapsed_seconds": 0.0,
                        "attempt": 0,
                        "fast_fail": True,
                        "systemic": True,
                        "status": "failed",
                        "error": (
                            "Skipped: circuit breaker after "
                            f"{consecutive_fast_fails[worker_index]} consecutive fast failures"
                        ),
                    }
                    for skip_item in remaining_items
                ]
                if skipped_failures:
                    apply_worker_results(worker_index, [], skipped_failures)
                    failures_count[worker_index] += len(skipped_failures)
                positions[worker_index] = len(group)
                continue

            if consecutive_failures[worker_index] >= settings.circuit_breaker_slow_fail_threshold:
                logger.error(
                    "IW51 worker CIRCUIT-BREAKER worker=%s consecutive_failures=%s — aborting worker, skipping %s remaining items.",
                    worker_index,
                    consecutive_failures[worker_index],
                    len(remaining_items),
                )
                worker_state.status = "circuit_breaker"
                skipped_failures = [
                    {
                        "worker_index": worker_index,
                        "row_index": skip_item.row_index,
                        "pn": skip_item.pn,
                        "instalacao": skip_item.instalacao,
                        "tipologia": skip_item.tipologia,
                        "elapsed_seconds": 0.0,
                        "attempt": 0,
                        "fast_fail": False,
                        "systemic": False,
                        "status": "failed",
                        "error": (
                            "Skipped: circuit breaker after "
                            f"{consecutive_failures[worker_index]} consecutive failures"
                        ),
                    }
                    for skip_item in remaining_items
                ]
                if skipped_failures:
                    apply_worker_results(worker_index, [], skipped_failures)
                    failures_count[worker_index] += len(skipped_failures)
                positions[worker_index] = len(group)

        if progressed:
            continue

        if len(completed_workers) < len(groups):
            pending_deadlines = [
                next_ready_at[worker_index]
                for worker_index in range(1, len(groups) + 1)
                if worker_index not in completed_workers and next_ready_at[worker_index] > time.monotonic()
            ]
            if pending_deadlines:
                time.sleep(max(0.0, min(pending_deadlines) - time.monotonic()))


def _run_iw51_parallel_worker(
    *,
    worker_index: int,
    session_locator: Iw51SessionLocator,
    items: list[Iw51WorkItem],
    settings: Iw51Settings,
    logger: Any,
    marshaled_app_stream: Any,
    worker_state: Iw51WorkerState,
    result_queue: queue.Queue[dict[str, Any]],
    cancel_event: threading.Event,
) -> None:
    worker_started_at = time.perf_counter()
    worker_deadline = time.monotonic() + settings.worker_timeout_seconds
    pythoncom, initialized = _co_initialize()
    session: Any | None = None
    application: Any | None = None
    next_item_index = 0
    try:
        worker_state.status = "running"
        worker_state.items_total = len(items)
        try:
            application = _unmarshal_sap_application(marshaled_app_stream, logger=logger, worker_index=worker_index)
        except Exception as exc:
            worker_state.status = "failed"
            logger.error(
                "IW51 worker ABORT worker=%s unmarshal failed — no items will be processed error=%s",
                worker_index,
                exc,
            )
            result_queue.put(
                {
                    "type": "results",
                    "worker_index": worker_index,
                    "results": [],
                    "failures": [
                        {
                            "worker_index": worker_index,
                            "row_index": item.row_index,
                            "pn": item.pn,
                            "instalacao": item.instalacao,
                            "tipologia": item.tipologia,
                            "elapsed_seconds": 0.0,
                            "attempt": 0,
                            "fast_fail": True,
                            "systemic": True,
                            "status": "failed",
                            "error": f"COM unmarshal failed: {exc}",
                        }
                        for item in items
                    ],
                }
            )
            return

        try:
            session = _reattach_iw51_session(locator=session_locator, logger=logger, application=application)
            _wait_session_ready(session, timeout_seconds=settings.per_step_timeout_seconds, logger=logger)
        except Exception as exc:
            worker_state.status = "failed"
            logger.error(
                "IW51 worker ABORT worker=%s initial session attach failed error=%s",
                worker_index,
                exc,
            )
            result_queue.put(
                {
                    "type": "results",
                    "worker_index": worker_index,
                    "results": [],
                    "failures": [
                        {
                            "worker_index": worker_index,
                            "row_index": item.row_index,
                            "pn": item.pn,
                            "instalacao": item.instalacao,
                            "tipologia": item.tipologia,
                            "elapsed_seconds": 0.0,
                            "attempt": 0,
                            "fast_fail": True,
                            "systemic": True,
                            "status": "failed",
                            "error": f"Session attach failed: {exc}",
                        }
                        for item in items
                    ],
                }
            )
            return

        worker_state.session_id = _read_session_id(session)
        logger.info(
            "IW51 worker START worker=%s assigned_items=%s con=%s ses=%s session_id=%s",
            worker_index,
            len(items),
            session_locator.connection_index,
            session_locator.session_index,
            worker_state.session_id,
        )

        consecutive_failures = 0
        consecutive_fast_fails = 0
        for next_item_index, item in enumerate(items, start=1):
            if cancel_event.is_set():
                remaining_items = items[next_item_index - 1 :]
                if remaining_items:
                    result_queue.put(
                        {
                            "type": "results",
                            "worker_index": worker_index,
                            "results": [],
                            "failures": [
                                {
                                    "worker_index": worker_index,
                                    "row_index": remaining_item.row_index,
                                    "pn": remaining_item.pn,
                                    "instalacao": remaining_item.instalacao,
                                    "tipologia": remaining_item.tipologia,
                                    "elapsed_seconds": 0.0,
                                    "attempt": 0,
                                    "fast_fail": True,
                                    "systemic": True,
                                    "status": "failed",
                                    "error": "Cancelled: all workers entered circuit breaker",
                                }
                                for remaining_item in remaining_items
                            ],
                        }
                    )
                worker_state.status = "cancelled"
                break
            if time.monotonic() > worker_deadline:
                remaining_items = items[next_item_index - 1 :]
                if remaining_items:
                    result_queue.put(
                        {
                            "type": "results",
                            "worker_index": worker_index,
                            "results": [],
                            "failures": [
                                {
                                    "worker_index": worker_index,
                                    "row_index": remaining_item.row_index,
                                    "pn": remaining_item.pn,
                                    "instalacao": remaining_item.instalacao,
                                    "tipologia": remaining_item.tipologia,
                                    "elapsed_seconds": 0.0,
                                    "attempt": 0,
                                    "fast_fail": True,
                                    "systemic": True,
                                    "status": "failed",
                                    "error": "Worker timeout exceeded",
                                }
                                for remaining_item in remaining_items
                            ],
                        }
                    )
                worker_state.status = "timeout"
                break
            if next_item_index == 1 or next_item_index % _IW51_PROGRESS_LOG_INTERVAL == 0:
                elapsed_so_far = time.perf_counter() - worker_started_at
                rate = worker_state.items_ok / elapsed_so_far if elapsed_so_far > 0 else 0.0
                logger.info(
                    "IW51 worker PROGRESS worker=%s index=%s/%s ok=%s fail=%s elapsed_s=%.1f rate=%.2f items/s",
                    worker_index,
                    next_item_index,
                    len(items),
                    worker_state.items_ok,
                    worker_state.items_failed,
                    elapsed_so_far,
                    rate,
                )

            worker_state.current_row_index = item.row_index
            if worker_state.first_item_started_at == 0.0:
                worker_state.first_item_started_at = time.time()
            result, failure, session = _process_iw51_item_with_retry(
                worker_index=worker_index,
                session=session,
                session_locator=session_locator,
                item=item,
                settings=settings,
                logger=logger,
                application=application,
                serialize_com_calls=False,
            )

            if result is not None:
                worker_state.items_processed += 1
                worker_state.items_ok += 1
                worker_state.last_ok_row_index = item.row_index
                worker_state.consecutive_failures = 0
                worker_state.last_item_finished_at = time.time()
                consecutive_failures = 0
                consecutive_fast_fails = 0
                result_queue.put(
                    {
                        "type": "results",
                        "worker_index": worker_index,
                        "results": [result],
                        "failures": [],
                    }
                )
                if next_item_index < len(items) and settings.inter_item_sleep_seconds > 0:
                    logger.info(
                        "IW51 worker SLEEP worker=%s seconds=%.1f after_row=%s",
                        worker_index,
                        settings.inter_item_sleep_seconds,
                        item.row_index,
                    )
                    time.sleep(settings.inter_item_sleep_seconds)
                continue

            assert failure is not None
            worker_state.items_processed += 1
            worker_state.items_failed += 1
            consecutive_failures += 1
            worker_state.consecutive_failures = consecutive_failures
            worker_state.last_item_finished_at = time.time()
            if failure.get("fast_fail", False):
                consecutive_fast_fails += 1
            else:
                consecutive_fast_fails = 0

            result_queue.put(
                {
                    "type": "results",
                    "worker_index": worker_index,
                    "results": [],
                    "failures": [failure],
                }
            )

            remaining_items = items[next_item_index:]
            if consecutive_fast_fails >= settings.circuit_breaker_fast_fail_threshold:
                logger.error(
                    "IW51 worker CIRCUIT-BREAKER worker=%s consecutive_fast_fails=%s — aborting worker, skipping %s remaining items.",
                    worker_index,
                    consecutive_fast_fails,
                    len(remaining_items),
                )
                worker_state.status = "circuit_breaker"
                skipped_failures = [
                    {
                        "worker_index": worker_index,
                        "row_index": skip_item.row_index,
                        "pn": skip_item.pn,
                        "instalacao": skip_item.instalacao,
                        "tipologia": skip_item.tipologia,
                        "elapsed_seconds": 0.0,
                        "attempt": 0,
                        "fast_fail": True,
                        "systemic": True,
                        "status": "failed",
                        "error": (
                            "Skipped: circuit breaker after "
                            f"{consecutive_fast_fails} consecutive fast failures"
                        ),
                    }
                    for skip_item in remaining_items
                ]
                if skipped_failures:
                    result_queue.put(
                        {
                            "type": "results",
                            "worker_index": worker_index,
                            "results": [],
                            "failures": skipped_failures,
                        }
                    )
                break

            if consecutive_failures >= settings.circuit_breaker_slow_fail_threshold:
                logger.error(
                    "IW51 worker CIRCUIT-BREAKER worker=%s consecutive_failures=%s — aborting worker, skipping %s remaining items.",
                    worker_index,
                    consecutive_failures,
                    len(remaining_items),
                )
                worker_state.status = "circuit_breaker"
                skipped_failures = [
                    {
                        "worker_index": worker_index,
                        "row_index": skip_item.row_index,
                        "pn": skip_item.pn,
                        "instalacao": skip_item.instalacao,
                        "tipologia": skip_item.tipologia,
                        "elapsed_seconds": 0.0,
                        "attempt": 0,
                        "fast_fail": False,
                        "systemic": False,
                        "status": "failed",
                        "error": (
                            "Skipped: circuit breaker after "
                            f"{consecutive_failures} consecutive failures"
                        ),
                    }
                    for skip_item in remaining_items
                ]
                if skipped_failures:
                    result_queue.put(
                        {
                            "type": "results",
                            "worker_index": worker_index,
                            "results": [],
                            "failures": skipped_failures,
                        }
                    )
                break
    except Exception as exc:
        worker_state.status = "failed"
        logger.exception("IW51 worker CRASHED worker=%s error=%s", worker_index, exc)
        remaining_items = items[max(0, next_item_index - 1) :]
        if remaining_items:
            result_queue.put(
                {
                    "type": "results",
                    "worker_index": worker_index,
                    "results": [],
                    "failures": [
                        {
                            "worker_index": worker_index,
                            "row_index": item.row_index,
                            "pn": item.pn,
                            "instalacao": item.instalacao,
                            "tipologia": item.tipologia,
                            "elapsed_seconds": 0.0,
                            "attempt": 0,
                            "fast_fail": True,
                            "systemic": True,
                            "status": "failed",
                            "error": f"Worker crashed: {exc}",
                        }
                        for item in remaining_items
                    ],
                }
            )
    finally:
        worker_elapsed = time.perf_counter() - worker_started_at
        if worker_state.status not in {"circuit_breaker", "failed", "cancelled", "timeout"}:
            worker_state.status = "completed"
        worker_state.current_row_index = 0
        worker_state.elapsed_seconds = round(worker_elapsed, 3)
        if worker_state.items_processed > 0:
            worker_state.avg_item_seconds = round(worker_elapsed / worker_state.items_processed, 3)
        logger.info(
            "IW51 worker DONE worker=%s ok=%s fail=%s total=%s elapsed_s=%.1f",
            worker_index,
            worker_state.items_ok,
            worker_state.items_failed,
            len(items),
            worker_elapsed,
        )
        result_queue.put({"type": "done", "worker_index": worker_index})
        if initialized and pythoncom is not None:
            pythoncom.CoUninitialize()


def _run_iw51_true_parallel_workers(
    *,
    groups: list[list[Iw51WorkItem]],
    session_locators: list[Iw51SessionLocator],
    settings: Iw51Settings,
    logger: Any,
    worker_states: list[Iw51WorkerState],
    apply_worker_results: Any,
) -> None:
    """Run one SAP-bound worker thread per session with a single main-thread collector.

    Thread safety contract:
    - Worker threads only communicate via result_queue.put().
    - The main thread is the sole consumer of result_queue.
    - apply_worker_results is called exclusively by the main thread.
    - No SAP COM proxy is shared across workers after unmarshal.
    """
    logger.info(
        "IW51 starting true parallel processing workers=%s total_items=%s",
        len(groups),
        sum(len(group) for group in groups),
    )
    marshaled_streams = _marshal_sap_application(len(groups), logger=logger)
    result_queue: queue.Queue[dict[str, Any]] = queue.Queue()
    cancel_event = threading.Event()
    threads: list[threading.Thread] = []

    for worker_index, group in enumerate(groups, start=1):
        thread = threading.Thread(
            target=_run_iw51_parallel_worker,
            kwargs={
                "worker_index": worker_index,
                "session_locator": session_locators[worker_index - 1],
                "items": group,
                "settings": settings,
                "logger": logger,
                "marshaled_app_stream": marshaled_streams[worker_index - 1],
                "worker_state": worker_states[worker_index - 1],
                "result_queue": result_queue,
                "cancel_event": cancel_event,
            },
            name=f"iw51-worker-{worker_index}",
            daemon=True,
        )
        thread.start()
        threads.append(thread)
        if worker_index < len(groups) and settings.worker_stagger_seconds > 0:
            time.sleep(settings.worker_stagger_seconds)

    completed_workers: set[int] = set()
    while len(completed_workers) < len(groups):
        message = result_queue.get()
        message_type = message.get("type")
        worker_index = int(message.get("worker_index", 0))
        if message_type == "results":
            apply_worker_results(
                worker_index,
                list(message.get("results", [])),
                list(message.get("failures", [])),
            )
            if not cancel_event.is_set() and worker_states and all(
                state.status in {"circuit_breaker", "failed", "cancelled", "timeout"} for state in worker_states
            ):
                cancel_event.set()
            continue
        if message_type == "done":
            completed_workers.add(worker_index)

    for thread in threads:
        thread.join()


class _Iw51QueueLogger:
    def __init__(self, *, worker_index: int, result_queue: Any) -> None:
        self.worker_index = worker_index
        self.result_queue = result_queue

    def _emit(self, level: str, message: str, *args: Any) -> None:
        formatted = message % args if args else message
        self.result_queue.put(
            {
                "type": "log",
                "worker_index": self.worker_index,
                "level": level,
                "message": formatted,
            }
        )

    def info(self, message: str, *args: Any) -> None:
        self._emit("info", message, *args)

    def warning(self, message: str, *args: Any) -> None:
        self._emit("warning", message, *args)

    def error(self, message: str, *args: Any) -> None:
        self._emit("error", message, *args)

    def exception(self, message: str, *args: Any) -> None:
        self._emit("error", message, *args)


def _iw51_process_heartbeat_loop(
    *,
    worker_index: int,
    result_queue: Any,
    stop_event: threading.Event,
    interval_seconds: float,
) -> None:
    while not stop_event.wait(max(1.0, interval_seconds)):
        result_queue.put(
            {
                "type": "heartbeat",
                "worker_index": worker_index,
                "sent_at": time.time(),
            }
        )


def _iw51_process_worker_main(
    *,
    worker_index: int,
    session_locator: Iw51SessionLocator,
    items: list[Iw51WorkItem],
    start_index: int,
    settings: Iw51Settings,
    result_queue: Any,
) -> None:
    logger = _Iw51QueueLogger(worker_index=worker_index, result_queue=result_queue)
    worker_state = Iw51WorkerState(worker_index=worker_index, session_id="")
    pythoncom, initialized = _co_initialize(apartment_threaded=True)
    previous_filter: Any | None = None
    wrapped_filter: Any | None = None
    heartbeat_stop = threading.Event()
    heartbeat_thread: threading.Thread | None = None
    session: Any | None = None
    application: Any | None = None
    local_index = 0
    try:
        previous_filter, wrapped_filter = _register_iw51_message_filter(
            retry_window_seconds=settings.message_filter_retry_window_seconds,
            logger=logger,
        )
        heartbeat_thread = threading.Thread(
            target=_iw51_process_heartbeat_loop,
            kwargs={
                "worker_index": worker_index,
                "result_queue": result_queue,
                "stop_event": heartbeat_stop,
                "interval_seconds": settings.worker_heartbeat_seconds,
            },
            name=f"iw51-process-heartbeat-{worker_index}",
            daemon=True,
        )
        heartbeat_thread.start()
        result_queue.put({"type": "heartbeat", "worker_index": worker_index, "sent_at": time.time()})

        application = _get_sap_application()
        session = _reattach_iw51_session(locator=session_locator, logger=logger, application=application)
        worker_state.session_id = _read_session_id(session) or ""
        _wait_session_ready(session, timeout_seconds=settings.per_step_timeout_seconds, logger=logger)

        for local_index, item in enumerate(items):
            result_queue.put(
                {
                    "type": "item_started",
                    "worker_index": worker_index,
                    "resume_index": start_index + local_index,
                    "row_index": item.row_index,
                    "sent_at": time.time(),
                }
            )
            result, failure, session = _process_iw51_item_with_retry(
                worker_index=worker_index,
                session=session,
                session_locator=session_locator,
                item=item,
                settings=settings,
                logger=logger,
                application=application,
                serialize_com_calls=False,
            )
            result_queue.put({"type": "heartbeat", "worker_index": worker_index, "sent_at": time.time()})
            if result is not None:
                result_queue.put(
                    {
                        "type": "results",
                        "worker_index": worker_index,
                        "results": [result],
                        "failures": [],
                    }
                )
                if local_index < len(items) - 1 and settings.inter_item_sleep_seconds > 0:
                    logger.info(
                        "IW51 process worker sleeping between items worker=%s seconds=%.1f",
                        worker_index,
                        settings.inter_item_sleep_seconds,
                    )
                    time.sleep(settings.inter_item_sleep_seconds)
                continue

            assert failure is not None
            if failure.get("systemic", False):
                result_queue.put(
                    {
                        "type": "needs_session_rebuild",
                        "worker_index": worker_index,
                        "resume_index": start_index + local_index,
                        "row_index": item.row_index,
                        "error": failure.get("error", ""),
                    }
                )
                return
            result_queue.put(
                {
                    "type": "results",
                    "worker_index": worker_index,
                    "results": [],
                    "failures": [failure],
                }
            )
            if local_index < len(items) - 1 and settings.inter_item_sleep_seconds > 0:
                logger.info(
                    "IW51 process worker sleeping between items worker=%s seconds=%.1f",
                    worker_index,
                    settings.inter_item_sleep_seconds,
                )
                time.sleep(settings.inter_item_sleep_seconds)

        result_queue.put(
            {
                "type": "done",
                "worker_index": worker_index,
                "resume_index": start_index + len(items),
            }
        )
    except Exception as exc:
        result_queue.put(
            {
                "type": "worker_fatal",
                "worker_index": worker_index,
                "resume_index": start_index + max(0, local_index),
                "error": str(exc),
            }
        )
    finally:
        heartbeat_stop.set()
        if heartbeat_thread is not None:
            heartbeat_thread.join(timeout=1.0)
        _unregister_iw51_message_filter(previous_filter, logger)
        if initialized and pythoncom is not None:
            pythoncom.CoUninitialize()


def _iw51_get_multiprocessing_context() -> Any:
    return multiprocessing.get_context("spawn")


def _start_iw51_process_worker(
    *,
    context: Any,
    result_queue: Any,
    worker_index: int,
    session_locator: Iw51SessionLocator,
    items: list[Iw51WorkItem],
    start_index: int,
    settings: Iw51Settings,
) -> Any:
    process = context.Process(
        target=_iw51_process_worker_main,
        kwargs={
            "worker_index": worker_index,
            "session_locator": session_locator,
            "items": items,
            "start_index": start_index,
            "settings": settings,
            "result_queue": result_queue,
        },
        name=f"iw51-process-worker-{worker_index}",
        daemon=True,
    )
    process.start()
    return process


def _stop_iw51_process_worker(process: Any | None) -> None:
    if process is None:
        return
    try:
        if process.is_alive():
            process.terminate()
    except Exception:
        return
    try:
        process.join(timeout=5.0)
    except Exception:
        return


def _update_iw51_worker_state_from_results(
    *,
    worker_state: Iw51WorkerState,
    worker_results: list[Iw51ItemResult],
    worker_failures: list[dict[str, Any]],
) -> None:
    processed_delta = len(worker_results) + len(worker_failures)
    worker_state.items_processed += processed_delta
    worker_state.items_ok += len(worker_results)
    worker_state.items_failed += len(worker_failures)
    worker_state.last_item_finished_at = time.time()
    if worker_results:
        worker_state.last_ok_row_index = worker_results[-1].row_index
        worker_state.consecutive_failures = 0
    elif worker_failures:
        worker_state.consecutive_failures += len(worker_failures)


def _run_iw51_group_in_main_process(
    *,
    worker_index: int,
    group: list[Iw51WorkItem],
    start_index: int,
    session: Any,
    session_locator: Iw51SessionLocator,
    settings: Iw51Settings,
    logger: Any,
    worker_state: Iw51WorkerState,
    apply_worker_results: Any,
) -> Any:
    current_session = session
    consecutive_failures = worker_state.consecutive_failures
    consecutive_fast_fails = 0
    for item_position, item in enumerate(group[start_index:], start=start_index):
        worker_state.status = "fallback_running"
        worker_state.current_row_index = item.row_index
        if worker_state.first_item_started_at == 0.0:
            worker_state.first_item_started_at = time.time()
        result, failure, current_session = _process_iw51_item_with_retry(
            worker_index=worker_index,
            session=current_session,
            session_locator=session_locator,
            item=item,
            settings=settings,
            logger=logger,
        )
        if result is not None:
            apply_worker_results(worker_index, [result], [])
            _update_iw51_worker_state_from_results(worker_state=worker_state, worker_results=[result], worker_failures=[])
            consecutive_failures = 0
            consecutive_fast_fails = 0
            if settings.inter_item_sleep_seconds > 0 and item_position < len(group) - 1:
                logger.info(
                    "Sleeping between IW51 items worker=%s seconds=%.1f",
                    worker_index,
                    settings.inter_item_sleep_seconds,
                )
                time.sleep(settings.inter_item_sleep_seconds)
            continue

        assert failure is not None
        apply_worker_results(worker_index, [], [failure])
        _update_iw51_worker_state_from_results(worker_state=worker_state, worker_results=[], worker_failures=[failure])
        consecutive_failures += 1
        worker_state.consecutive_failures = consecutive_failures
        if failure.get("fast_fail", False):
            consecutive_fast_fails += 1
        else:
            consecutive_fast_fails = 0
        if consecutive_fast_fails >= settings.circuit_breaker_fast_fail_threshold or consecutive_failures >= settings.circuit_breaker_slow_fail_threshold:
            remaining_items = group[group.index(item) + 1 :]
            skipped_failures = [
                {
                    "worker_index": worker_index,
                    "row_index": skip_item.row_index,
                    "pn": skip_item.pn,
                    "instalacao": skip_item.instalacao,
                    "tipologia": skip_item.tipologia,
                    "elapsed_seconds": 0.0,
                    "attempt": 0,
                    "fast_fail": True,
                    "systemic": True,
                    "status": "failed",
                    "error": "Skipped: fallback circuit breaker",
                }
                for skip_item in remaining_items
            ]
            if skipped_failures:
                apply_worker_results(worker_index, [], skipped_failures)
                _update_iw51_worker_state_from_results(
                    worker_state=worker_state,
                    worker_results=[],
                    worker_failures=skipped_failures,
                )
            worker_state.status = "circuit_breaker"
            worker_state.current_row_index = 0
            return current_session
        if settings.inter_item_sleep_seconds > 0 and item_position < len(group) - 1:
            logger.info(
                "Sleeping between IW51 items worker=%s seconds=%.1f",
                worker_index,
                settings.inter_item_sleep_seconds,
            )
            time.sleep(settings.inter_item_sleep_seconds)
    worker_state.status = "completed"
    worker_state.current_row_index = 0
    return current_session


def _rebuild_iw51_worker_session(
    *,
    worker_index: int,
    locator: Iw51SessionLocator,
    settings: Iw51Settings,
    logger: Any,
) -> Any:
    if settings.session_rebuild_backoff_seconds > 0:
        time.sleep(settings.session_rebuild_backoff_seconds)
    session = _reattach_iw51_session(locator=locator, logger=logger)
    _wait_session_ready(session, timeout_seconds=settings.per_step_timeout_seconds, logger=logger)
    if settings.session_recovery_mode == "soft":
        _reset_session_state_soft(session=session, settings=settings, logger=logger)
    logger.info(
        "IW51 process worker session rebuilt worker=%s con=%s ses=%s session_id=%s",
        worker_index,
        locator.connection_index,
        locator.session_index,
        _read_session_id(session) or "<empty>",
    )
    return session


def _run_iw51_process_parallel_workers(
    *,
    groups: list[list[Iw51WorkItem]],
    session_locators: list[Iw51SessionLocator],
    sessions: list[Any],
    settings: Iw51Settings,
    logger: Any,
    worker_states: list[Iw51WorkerState],
    apply_worker_results: Any,
) -> tuple[dict[str, int], dict[str, int], dict[str, float]]:
    logger.info(
        "IW51 starting process parallel processing workers=%s total_items=%s",
        len(groups),
        sum(len(group) for group in groups),
    )
    context = _iw51_get_multiprocessing_context()
    result_queue = context.Queue()
    runtimes: dict[int, _Iw51ProcessWorkerRuntime] = {}

    for worker_index, group in enumerate(groups, start=1):
        runtime = _Iw51ProcessWorkerRuntime(
            worker_index=worker_index,
            locator=session_locators[worker_index - 1],
            group=group,
            last_heartbeat_at=time.monotonic(),
        )
        runtime.process = _start_iw51_process_worker(
            context=context,
            result_queue=result_queue,
            worker_index=worker_index,
            session_locator=runtime.locator,
            items=group,
            start_index=0,
            settings=settings,
        )
        runtimes[worker_index] = runtime
        worker_states[worker_index - 1].status = "running"
        worker_states[worker_index - 1].items_total = len(group)
        if worker_index < len(groups) and settings.worker_start_stagger_seconds > 0:
            time.sleep(settings.worker_start_stagger_seconds)

    completed_workers = 0
    try:
        while completed_workers < len(groups):
            now = time.monotonic()
            try:
                message = result_queue.get(timeout=0.5)
            except queue.Empty:
                message = None

            if message is not None:
                worker_index = int(message.get("worker_index", 0))
                runtime = runtimes[worker_index]
                worker_state = worker_states[worker_index - 1]
                message_type = str(message.get("type", ""))
                if message_type == "log":
                    log_level = str(message.get("level", "info")).lower()
                    log_fn = getattr(logger, log_level, logger.info)
                    log_fn("%s", message.get("message", ""))
                elif message_type == "heartbeat":
                    runtime.last_heartbeat_at = now
                elif message_type == "item_started":
                    runtime.last_heartbeat_at = now
                    runtime.next_item_index = int(message.get("resume_index", runtime.next_item_index))
                    runtime.current_row_index = int(message.get("row_index", 0))
                    runtime.current_item_started_at = now
                    worker_state.current_row_index = runtime.current_row_index
                    if worker_state.first_item_started_at == 0.0:
                        worker_state.first_item_started_at = time.time()
                elif message_type == "results":
                    worker_results = list(message.get("results", []))
                    worker_failures = list(message.get("failures", []))
                    runtime.last_heartbeat_at = now
                    runtime.current_item_started_at = 0.0
                    runtime.current_row_index = 0
                    worker_state.current_row_index = 0
                    apply_worker_results(worker_index, worker_results, worker_failures)
                    _update_iw51_worker_state_from_results(
                        worker_state=worker_state,
                        worker_results=worker_results,
                        worker_failures=worker_failures,
                    )
                    runtime.next_item_index += len(worker_results) + len(worker_failures)
                elif message_type in {"needs_session_rebuild", "worker_fatal"}:
                    runtime.last_heartbeat_at = now
                    runtime.current_item_started_at = 0.0
                    runtime.next_item_index = int(message.get("resume_index", runtime.next_item_index))
                    runtime.current_row_index = int(message.get("row_index", runtime.current_row_index))
                    worker_state.current_row_index = runtime.current_row_index
                    _stop_iw51_process_worker(runtime.process)
                    if message_type == "needs_session_rebuild":
                        runtime.session_rebuild_count += 1
                        logger.warning(
                            "IW51 process worker requested session rebuild worker=%s row=%s error=%s",
                            worker_index,
                            runtime.current_row_index or "<empty>",
                            message.get("error", ""),
                        )
                    else:
                        logger.warning(
                            "IW51 process worker fatal worker=%s resume_index=%s error=%s",
                            worker_index,
                            runtime.next_item_index,
                            message.get("error", ""),
                        )
                    if runtime.restart_count >= settings.worker_restart_limit:
                        logger.error(
                            "IW51 process worker exceeded restart limit worker=%s limit=%s — falling back to main process",
                            worker_index,
                            settings.worker_restart_limit,
                        )
                        sessions[worker_index - 1] = _rebuild_iw51_worker_session(
                            worker_index=worker_index,
                            locator=runtime.locator,
                            settings=settings,
                            logger=logger,
                        )
                        sessions[worker_index - 1] = _run_iw51_group_in_main_process(
                            worker_index=worker_index,
                            group=runtime.group,
                            start_index=runtime.next_item_index,
                            session=sessions[worker_index - 1],
                            session_locator=runtime.locator,
                            settings=settings,
                            logger=logger,
                            worker_state=worker_state,
                            apply_worker_results=apply_worker_results,
                        )
                        runtime.done = True
                        completed_workers += 1
                        continue
                    runtime.restart_count += 1
                    if settings.worker_restart_backoff_seconds > 0:
                        time.sleep(settings.worker_restart_backoff_seconds)
                    sessions[worker_index - 1] = _rebuild_iw51_worker_session(
                        worker_index=worker_index,
                        locator=runtime.locator,
                        settings=settings,
                        logger=logger,
                    )
                    runtime.process = _start_iw51_process_worker(
                        context=context,
                        result_queue=result_queue,
                        worker_index=worker_index,
                        session_locator=runtime.locator,
                        items=runtime.group[runtime.next_item_index :],
                        start_index=runtime.next_item_index,
                        settings=settings,
                    )
                    worker_state.status = "running"
                elif message_type == "done" and not runtime.done:
                    runtime.done = True
                    runtime.current_item_started_at = 0.0
                    runtime.current_row_index = 0
                    worker_state.current_row_index = 0
                    if worker_state.status not in {"circuit_breaker", "failed", "fallback_running"}:
                        worker_state.status = "completed"
                    completed_workers += 1

            for worker_index, runtime in runtimes.items():
                if runtime.done:
                    continue
                worker_state = worker_states[worker_index - 1]
                if runtime.current_item_started_at and (now - runtime.current_item_started_at) > settings.worker_item_timeout_seconds:
                    logger.error(
                        "IW51 process worker item timeout worker=%s row=%s elapsed_s=%.1f",
                        worker_index,
                        runtime.current_row_index or "<empty>",
                        now - runtime.current_item_started_at,
                    )
                    _stop_iw51_process_worker(runtime.process)
                    if runtime.restart_count >= settings.worker_restart_limit:
                        logger.error(
                            "IW51 process worker timeout exceeded restart limit worker=%s — falling back to main process",
                            worker_index,
                        )
                        sessions[worker_index - 1] = _rebuild_iw51_worker_session(
                            worker_index=worker_index,
                            locator=runtime.locator,
                            settings=settings,
                            logger=logger,
                        )
                        sessions[worker_index - 1] = _run_iw51_group_in_main_process(
                            worker_index=worker_index,
                            group=runtime.group,
                            start_index=runtime.next_item_index,
                            session=sessions[worker_index - 1],
                            session_locator=runtime.locator,
                            settings=settings,
                            logger=logger,
                            worker_state=worker_state,
                            apply_worker_results=apply_worker_results,
                        )
                        runtime.done = True
                        completed_workers += 1
                        continue
                    runtime.restart_count += 1
                    runtime.session_rebuild_count += 1
                    sessions[worker_index - 1] = _rebuild_iw51_worker_session(
                        worker_index=worker_index,
                        locator=runtime.locator,
                        settings=settings,
                        logger=logger,
                    )
                    runtime.process = _start_iw51_process_worker(
                        context=context,
                        result_queue=result_queue,
                        worker_index=worker_index,
                        session_locator=runtime.locator,
                        items=runtime.group[runtime.next_item_index :],
                        start_index=runtime.next_item_index,
                        settings=settings,
                    )
                    runtime.current_item_started_at = 0.0
                    worker_state.status = "running"
                    continue
                if runtime.process is not None and getattr(runtime.process, "exitcode", None) not in {None, 0} and not runtime.done:
                    logger.warning(
                        "IW51 process worker exited unexpectedly worker=%s exitcode=%s",
                        worker_index,
                        getattr(runtime.process, "exitcode", None),
                    )
                    _stop_iw51_process_worker(runtime.process)
                    if runtime.restart_count >= settings.worker_restart_limit:
                        logger.error(
                            "IW51 process worker unexpected exit exceeded restart limit worker=%s — falling back to main process",
                            worker_index,
                        )
                        sessions[worker_index - 1] = _rebuild_iw51_worker_session(
                            worker_index=worker_index,
                            locator=runtime.locator,
                            settings=settings,
                            logger=logger,
                        )
                        sessions[worker_index - 1] = _run_iw51_group_in_main_process(
                            worker_index=worker_index,
                            group=runtime.group,
                            start_index=runtime.next_item_index,
                            session=sessions[worker_index - 1],
                            session_locator=runtime.locator,
                            settings=settings,
                            logger=logger,
                            worker_state=worker_state,
                            apply_worker_results=apply_worker_results,
                        )
                        runtime.done = True
                        completed_workers += 1
                        continue
                    runtime.restart_count += 1
                    sessions[worker_index - 1] = _rebuild_iw51_worker_session(
                        worker_index=worker_index,
                        locator=runtime.locator,
                        settings=settings,
                        logger=logger,
                    )
                    runtime.process = _start_iw51_process_worker(
                        context=context,
                        result_queue=result_queue,
                        worker_index=worker_index,
                        session_locator=runtime.locator,
                        items=runtime.group[runtime.next_item_index :],
                        start_index=runtime.next_item_index,
                        settings=settings,
                    )
                    worker_state.status = "running"

    finally:
        for runtime in runtimes.values():
            _stop_iw51_process_worker(runtime.process)

    return (
        {str(worker_index): runtime.restart_count for worker_index, runtime in runtimes.items()},
        {str(worker_index): runtime.session_rebuild_count for worker_index, runtime in runtimes.items()},
        {
            str(worker_index): round(max(0.0, time.monotonic() - runtime.last_heartbeat_at), 3)
            for worker_index, runtime in runtimes.items()
        },
    )


def run_iw51_demandante(
    *,
    run_id: str,
    demandante: str,
    config_path: Path,
    output_root: Path,
    max_rows: int | None = None,
) -> Iw51Manifest:
    from .service import create_session_provider

    config = load_export_config(config_path)
    settings = load_iw51_settings(
        config=config,
        config_path=config_path,
        demandante=demandante,
    )
    logger, log_path = configure_run_logger(output_root=output_root, run_id=run_id)

    run_root = output_root.expanduser().resolve() / "runs" / run_id / "iw51"
    manifest_path = run_root / "iw51_manifest.json"
    working_workbook_path = run_root / "working" / settings.workbook_path.name
    ledger_path = run_root / "iw51_progress.csv"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info(
        "Starting IW51 run run_id=%s demandante=%s source_workbook=%s sheet=%s session_count=%s execution_mode=%s",
        run_id,
        settings.demandante,
        settings.workbook_path,
        settings.sheet_name,
        settings.session_count,
        settings.execution_mode,
    )

    _copy_working_workbook(
        source_path=settings.workbook_path,
        working_path=working_workbook_path,
        logger=logger,
    )
    recovered_from_log = False
    if not ledger_path.exists() and log_path.exists():
        try:
            recovery_summary = recover_iw51_run_artifacts_from_log(
                run_root=run_root.parent,
                workbook_path=working_workbook_path if working_workbook_path.exists() else None,
                sheet_name=settings.sheet_name,
            )
            recovered_from_log = recovery_summary.get("recovered_success_rows", 0) > 0
            if recovered_from_log:
                logger.info(
                    "IW51 ledger rebuilt from log recovered_success_rows=%s log=%s",
                    recovery_summary.get("recovered_success_rows", 0),
                    recovery_summary.get("log_path", log_path),
                )
        except Exception as exc:
            logger.warning("IW51 log recovery skipped error=%s", exc)
    ledger_state = _load_iw51_ledger_state(ledger_path)
    if ledger_state.total_entries > 0 and ledger_state.total_entries > ledger_state.unique_rows * 2:
        removed_entries = _compact_iw51_ledger(ledger_path)
        if removed_entries > 0:
            logger.info(
                "IW51 ledger compacted removed_entries=%s ledger=%s",
                removed_entries,
                ledger_path,
            )
            ledger_state = _load_iw51_ledger_state(ledger_path)
    ledger_success_rows = set(ledger_state.success_rows)
    ledger_terminal_rows = set(ledger_state.terminal_rows)

    workbook, sheet, feito_column_index, items, skipped_rows, rejected_failures = load_iw51_work_items(
        workbook_path=working_workbook_path,
        sheet_name=settings.sheet_name,
        max_rows=max_rows or settings.max_rows_per_run,
        completed_row_indices=ledger_terminal_rows,
    )
    header_index_by_key, existing_feito_column_index, data_start_row = _resolve_iw51_sheet_layout(sheet)
    workbook_done_rows = _collect_iw51_workbook_done_rows(
        sheet=sheet,
        header_index_by_key=header_index_by_key,
        feito_column_index=feito_column_index or existing_feito_column_index,
        data_start_row=data_start_row,
        ledger_terminal_rows=ledger_terminal_rows,
    )
    if workbook_done_rows:
        append_iw51_progress_ledger(ledger_path=ledger_path, rows=workbook_done_rows)
        imported_rows = {int(row["row_index"]) for row in workbook_done_rows}
        ledger_success_rows.update(imported_rows)
        ledger_terminal_rows.update(imported_rows)
        logger.info(
            "IW51 ledger reconciled from workbook feito_rows=%s ledger=%s",
            len(workbook_done_rows),
            ledger_path,
        )

    if ledger_success_rows:
        _sync_workbook_done_rows(
            workbook=workbook,
            sheet=sheet,
            feito_column_index=feito_column_index,
            workbook_path=working_workbook_path,
            row_indices=ledger_success_rows,
            logger=logger,
        )
        logger.info(
            "IW51 ledger reconciled into workbook rows=%s workbook=%s",
            len(ledger_success_rows),
            working_workbook_path,
        )

    workbook_done_count = _count_iw51_workbook_done_rows(
        sheet=sheet,
        feito_column_index=feito_column_index,
        data_start_row=data_start_row,
    )
    logger.info(
        "IW51 RERUN SUMMARY run_id=%s ledger_success=%s ledger_rejected=%s ledger_failed_terminal=%s ledger_failed_retriable=%s workbook_feito=%s pending_items=%s skipped=%s",
        run_id,
        len(ledger_success_rows),
        len(ledger_state.rejected_rows),
        len(ledger_state.failed_terminal_rows),
        len(ledger_state.retriable_failed_rows),
        workbook_done_count,
        len(items),
        skipped_rows,
    )
    if workbook_done_count != len(ledger_success_rows):
        logger.warning(
            "IW51 reconciliation mismatch ledger_success=%s workbook_feito=%s workbook=%s ledger=%s",
            len(ledger_success_rows),
            workbook_done_count,
            working_workbook_path,
            ledger_path,
        )

    failed_rows: list[dict[str, Any]] = list(rejected_failures)
    successful_rows = 0
    rejected_rows = len(rejected_failures)
    pending_sync_rows: set[int] = set()
    worker_states: list[Iw51WorkerState] = []
    worker_restarts: dict[str, int] = {}
    session_rebuilds: dict[str, int] = {}
    heartbeat_lag_seconds: dict[str, float] = {}
    collector_thread = threading.current_thread()
    run_started_at = time.perf_counter()
    total_items_target = len(items) + len(rejected_failures)

    if rejected_failures:
        append_iw51_progress_ledger(
            ledger_path=ledger_path,
            rows=[_build_ledger_row_from_failure(row) for row in rejected_failures],
        )

    if not items:
        manifest = Iw51Manifest(
            run_id=run_id,
            demandante=settings.demandante,
            workbook_path=str(settings.workbook_path),
            source_workbook_path=str(settings.workbook_path),
            working_workbook_path=str(working_workbook_path),
            progress_ledger_path=str(ledger_path),
            sheet_name=settings.sheet_name,
            processed_rows=successful_rows + len(failed_rows),
            successful_rows=successful_rows,
            skipped_rows=skipped_rows,
            rejected_rows=rejected_rows,
            status="skipped" if not failed_rows else "partial",
            log_path=str(log_path),
            manifest_path=str(manifest_path),
            supervisor_mode=settings.execution_mode,
            recovered_from_log=recovered_from_log,
            worker_restarts=worker_restarts,
            session_rebuilds=session_rebuilds,
            heartbeat_lag_seconds=heartbeat_lag_seconds,
            remaining_rows=0,
            failed_rows=failed_rows,
            worker_states=[],
        )
        manifest_path.write_text(json.dumps(manifest.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
        return manifest

    session_provider = create_session_provider(config=config)
    base_session = session_provider.get_session(config=config, logger=logger)
    sessions = prepare_iw51_sessions(
        base_session=base_session,
        settings=settings,
        logger=logger,
    )
    session_locators = [_session_locator_from_session(session) for session in sessions]
    groups = split_iw51_work_items_evenly(items, len(sessions))

    for worker_index, group in enumerate(groups, start=1):
        logger.info(
            "IW51 assigned instruction batch slot=%s items=%s first_row=%s last_row=%s",
            worker_index,
            len(group),
            group[0].row_index if group else "<none>",
            group[-1].row_index if group else "<none>",
        )
    worker_states = [
        Iw51WorkerState(
            worker_index=worker_index,
            session_id=f"/app/con[{session_locators[worker_index - 1].connection_index}]/ses[{session_locators[worker_index - 1].session_index}]",
        )
        for worker_index in range(1, len(groups) + 1)
    ]

    def _persist_progress() -> None:
        nonlocal workbook, sheet
        if len(pending_sync_rows) >= settings.workbook_sync_batch_size:
            _flush_pending_sync_rows()

    def _flush_pending_sync_rows() -> None:
        nonlocal workbook, sheet
        if not pending_sync_rows:
            return
        rows_to_sync = set(pending_sync_rows)
        pending_sync_rows.clear()
        _sync_workbook_done_rows(
            workbook=workbook,
            sheet=sheet,
            feito_column_index=feito_column_index,
            workbook_path=working_workbook_path,
            row_indices=rows_to_sync,
            logger=logger,
        )

    def _apply_worker_results(
        worker_index: int,
        worker_results: list[Iw51ItemResult],
        worker_failures: list[dict[str, Any]],
    ) -> None:
        assert threading.current_thread() is collector_thread, (
            "_apply_worker_results must be called from the collector thread, "
            f"got {threading.current_thread().name}"
        )
        nonlocal successful_rows
        ledger_rows: list[dict[str, Any]] = []
        for result in worker_results:
            successful_rows += 1
            pending_sync_rows.add(result.row_index)
            ledger_rows.append(_build_ledger_row_from_result(result))
        for failure in worker_failures:
            failed_rows.append(failure)
            ledger_rows.append(_build_ledger_row_from_failure(failure))
        append_iw51_progress_ledger(ledger_path=ledger_path, rows=ledger_rows)
        _persist_progress()
        total_collected = successful_rows + len(failed_rows)
        manifest = Iw51Manifest(
            run_id=run_id,
            demandante=settings.demandante,
            workbook_path=str(settings.workbook_path),
            source_workbook_path=str(settings.workbook_path),
            working_workbook_path=str(working_workbook_path),
            progress_ledger_path=str(ledger_path),
            sheet_name=settings.sheet_name,
            processed_rows=successful_rows + len(failed_rows),
            successful_rows=successful_rows,
            skipped_rows=skipped_rows,
            rejected_rows=rejected_rows,
            status="success"
            if successful_rows == len(items) and not failed_rows
            else "partial",
            log_path=str(log_path),
            manifest_path=str(manifest_path),
            supervisor_mode=settings.execution_mode,
            recovered_from_log=recovered_from_log,
            worker_restarts=worker_restarts,
            session_rebuilds=session_rebuilds,
            heartbeat_lag_seconds=heartbeat_lag_seconds,
            remaining_rows=max(0, total_items_target - total_collected),
            failed_rows=failed_rows,
            worker_states=[state.to_dict() for state in worker_states],
        )
        manifest_path.write_text(json.dumps(manifest.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
        logger.info(
            "IW51 worker COLLECTED worker=%s successful=%s failures=%s",
            worker_index,
            len(worker_results),
            len(worker_failures),
        )
        if total_collected > 0 and (total_collected % 50 == 0 or total_collected == total_items_target):
            elapsed_seconds = time.perf_counter() - run_started_at
            rate = total_collected / elapsed_seconds if elapsed_seconds > 0 else 0.0
            eta_seconds = ((total_items_target - total_collected) / rate) if rate > 0 else float("inf")
            logger.info(
                "IW51 PROGRESS total=%s/%s ok=%s fail=%s elapsed_s=%.1f rate=%.2f items/s eta_s=%s",
                total_collected,
                total_items_target,
                successful_rows,
                len(failed_rows),
                elapsed_seconds,
                rate,
                "inf" if eta_seconds == float("inf") else f"{eta_seconds:.0f}",
            )

    try:
        if settings.execution_mode == "process_parallel" and len(groups) > 1:
            worker_restarts, session_rebuilds, heartbeat_lag_seconds = _run_iw51_process_parallel_workers(
                groups=groups,
                session_locators=session_locators,
                sessions=sessions,
                settings=settings,
                logger=logger,
                worker_states=worker_states,
                apply_worker_results=_apply_worker_results,
            )
        elif settings.execution_mode == "true_parallel" and len(groups) > 1:
            try:
                _run_iw51_true_parallel_workers(
                    groups=groups,
                    session_locators=session_locators,
                    settings=settings,
                    logger=logger,
                    worker_states=worker_states,
                    apply_worker_results=_apply_worker_results,
                )
            except Exception as exc:
                logger.exception(
                    "IW51 true parallel processing crashed; falling back to interleaved mode error=%s",
                    exc,
                )
                for worker_state in worker_states:
                    if worker_state.status not in {"completed", "circuit_breaker"}:
                        worker_state.status = "idle"
                        worker_state.current_row_index = 0
                        worker_state.consecutive_failures = 0
                _run_iw51_interleaved_workers(
                    groups=groups,
                    session_locators=session_locators,
                    sessions=sessions,
                    settings=Iw51Settings(
                        demandante=settings.demandante,
                        workbook_path=settings.workbook_path,
                        sheet_name=settings.sheet_name,
                        inter_item_sleep_seconds=settings.inter_item_sleep_seconds,
                        max_rows_per_run=settings.max_rows_per_run,
                        notification_type=settings.notification_type,
                        reference_notification_number=settings.reference_notification_number,
                        session_count=settings.session_count,
                        execution_mode="interleaved",
                        post_login_wait_seconds=settings.post_login_wait_seconds,
                        workbook_sync_batch_size=settings.workbook_sync_batch_size,
                        circuit_breaker_fast_fail_threshold=settings.circuit_breaker_fast_fail_threshold,
                        circuit_breaker_slow_fail_threshold=settings.circuit_breaker_slow_fail_threshold,
                        per_step_timeout_seconds=settings.per_step_timeout_seconds,
                        session_recovery_mode=settings.session_recovery_mode,
                        worker_stagger_seconds=settings.worker_stagger_seconds,
                        worker_timeout_seconds=settings.worker_timeout_seconds,
                    ),
                    logger=logger,
                    worker_states=worker_states,
                    apply_worker_results=_apply_worker_results,
                )
        elif settings.execution_mode == "interleaved" and len(groups) > 1:
            _run_iw51_interleaved_workers(
                groups=groups,
                session_locators=session_locators,
                sessions=sessions,
                settings=settings,
                logger=logger,
                worker_states=worker_states,
                apply_worker_results=_apply_worker_results,
            )
        else:
            logger.info("IW51 starting sequential processing workers=%s total_items=%s", len(groups), len(items))
            for worker_index, group in enumerate(groups, start=1):
                worker_state = worker_states[worker_index - 1]
                worker_state.status = "running"
                worker_state.items_total = len(group)
                current_session = sessions[worker_index - 1]
                started_at = time.perf_counter()
                for index, item in enumerate(group, start=1):
                    worker_state.current_row_index = item.row_index
                    if worker_state.first_item_started_at == 0.0:
                        worker_state.first_item_started_at = time.time()
                    result, failure, current_session = _process_iw51_item_with_retry(
                        worker_index=worker_index,
                        session=current_session,
                        session_locator=session_locators[worker_index - 1],
                        item=item,
                        settings=settings,
                        logger=logger,
                    )
                    if result is not None:
                        worker_state.items_processed += 1
                        worker_state.items_ok += 1
                        worker_state.last_ok_row_index = item.row_index
                        worker_state.last_item_finished_at = time.time()
                        _apply_worker_results(worker_index, [result], [])
                        if index < len(group) and settings.inter_item_sleep_seconds > 0:
                            logger.info(
                                "Sleeping between IW51 items worker=%s seconds=%.1f",
                                worker_index,
                                settings.inter_item_sleep_seconds,
                            )
                            time.sleep(settings.inter_item_sleep_seconds)
                        continue

                    assert failure is not None
                    worker_state.items_processed += 1
                    worker_state.items_failed += 1
                    worker_state.consecutive_failures += 1
                    worker_state.last_item_finished_at = time.time()
                    _apply_worker_results(worker_index, [], [failure])
                    if worker_state.consecutive_failures >= settings.circuit_breaker_slow_fail_threshold:
                        worker_state.status = "circuit_breaker"
                        remaining_items = group[index:]
                        skipped_failures = [
                            {
                                "worker_index": worker_index,
                                "row_index": skip_item.row_index,
                                "pn": skip_item.pn,
                                "instalacao": skip_item.instalacao,
                                "tipologia": skip_item.tipologia,
                                "elapsed_seconds": 0.0,
                                "attempt": 0,
                                "fast_fail": False,
                                "systemic": False,
                                "status": "failed",
                                "error": (
                                    "Skipped: circuit breaker after "
                                    f"{worker_state.consecutive_failures} consecutive failures"
                                ),
                            }
                            for skip_item in remaining_items
                        ]
                        if skipped_failures:
                            _apply_worker_results(worker_index, [], skipped_failures)
                        break
                if worker_state.status not in {"circuit_breaker", "failed"}:
                    worker_state.status = "completed"
                worker_state.elapsed_seconds = round(time.perf_counter() - started_at, 3)
                if worker_state.items_processed > 0:
                    worker_state.avg_item_seconds = round(worker_state.elapsed_seconds / worker_state.items_processed, 3)
                worker_state.current_row_index = 0
    finally:
        _flush_pending_sync_rows()

    status = (
        "skipped"
        if not items and not failed_rows and successful_rows == 0
        else "success"
        if successful_rows == len(items) and not failed_rows
        else "partial"
    )
    manifest = Iw51Manifest(
        run_id=run_id,
        demandante=settings.demandante,
        workbook_path=str(settings.workbook_path),
        source_workbook_path=str(settings.workbook_path),
        working_workbook_path=str(working_workbook_path),
        progress_ledger_path=str(ledger_path),
        sheet_name=settings.sheet_name,
        processed_rows=successful_rows + len(failed_rows),
        successful_rows=successful_rows,
        skipped_rows=skipped_rows,
        rejected_rows=rejected_rows,
        status=status,
        log_path=str(log_path),
        manifest_path=str(manifest_path),
        supervisor_mode=settings.execution_mode,
        recovered_from_log=recovered_from_log,
        worker_restarts=worker_restarts,
        session_rebuilds=session_rebuilds,
        heartbeat_lag_seconds=heartbeat_lag_seconds,
        remaining_rows=max(0, total_items_target - (successful_rows + len(failed_rows))),
        failed_rows=failed_rows,
        worker_states=[state.to_dict() for state in worker_states],
    )
    manifest_path.write_text(json.dumps(manifest.to_dict(), ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info(
        "IW51 run finished demandante=%s status=%s successful_rows=%s processed_rows=%s manifest=%s",
        settings.demandante,
        status,
        successful_rows,
        manifest.processed_rows,
        manifest_path,
    )
    return manifest


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run IW51 SAP GUI automation for a demandante workbook flow.")
    parser.add_argument("--run-id", required=True, help="Run identifier.")
    parser.add_argument("--demandante", default="DANI", help="Demandante profile. Default: DANI.")
    parser.add_argument("--config", default="sap_iw69_batch_config.json", help="Path to config JSON.")
    parser.add_argument("--output-root", default="output", help="Root folder for run artifacts.")
    parser.add_argument(
        "--max-rows",
        type=int,
        default=0,
        help="Maximum number of pending workbook rows to process. Default uses config.",
    )
    return parser


def main() -> int:
    args = build_arg_parser().parse_args()
    manifest = run_iw51_demandante(
        run_id=args.run_id,
        demandante=args.demandante,
        config_path=Path(args.config),
        output_root=Path(args.output_root),
        max_rows=args.max_rows or None,
    )
    return 0 if manifest.status in {"success", "skipped"} else 1


if __name__ == "__main__":
    raise SystemExit(main())
